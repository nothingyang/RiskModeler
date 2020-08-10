import tkinter as tk
from tkinter import *
import tkinter.messagebox
from tkinter import ttk
from tkinter.scrolledtext import *
from sklearn.metrics import roc_curve, auc
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import PatternFill,  colors
from openpyxl.formatting.rule import ColorScaleRule
import math
mpl.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['font.sans-serif']=['SimHei']
mpl.rcParams['axes.unicode_minus']=False
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
class ScrollableFrame(ttk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        canvas = tk.Canvas(self)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        hscrollbar = ttk.Scrollbar(self, orient=tk.HORIZONTAL, command=canvas.xview)

        self.scrollable_frame = ttk.Frame(canvas)
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
        def _bound_to_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)  

        def _unbound_to_mousewheel( event):
            canvas.bind_all("<MouseWheel>") 
#         canvas.bind('<Enter>', _unbound_to_mousewheel)
#         canvas.bind('<Leave>', _bound_to_mousewheel)

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set,xscrollcommand=hscrollbar.set)
        hscrollbar.pack(side="bottom", fill="x")
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        


class scorecard_result_ui():
    def __init__(self,mainframe,project_path,node_name,
                 predict_train_data,predict_vaild_data,predict_oot_data,predict_reject_data,
                 train_target,oot_target,reject_target,
                 train_time_id,oot_time_id,reject_time_id,
                 record_list, model,scorecarddf,f_group_report,variable_list,lasso_df,model_var_type,var_clus
                ):
        self.master=mainframe
        self.project_path =project_path
        self.node_name = node_name
        self.predict_train_data=predict_train_data
        self.predict_vaild_data=predict_vaild_data
        self.predict_oot_data=predict_oot_data
        self.predict_reject_data=predict_reject_data
        self.train_target=train_target
        self.oot_target=oot_target
        self.reject_target=reject_target
        self.train_time_id=train_time_id
        self.oot_time_id=oot_time_id
        self.reject_time_id=reject_time_id
        self.record_list =record_list
        self.model=model
        self.variable_list=variable_list
        self.scorecarddf=scorecarddf
        self.f_group_report=f_group_report
        self.model_var_type=model_var_type
        self.lasso_df=lasso_df
        self.var_clus=var_clus
        self.plot_tab(
                 predict_train_data=self.predict_train_data,predict_vaild_data=self.predict_vaild_data,
                 predict_oot_data=self.predict_oot_data,predict_reject_data=self.predict_reject_data,
                 train_target=self.train_target,oot_target=self.oot_target,reject_target=self.reject_target,
                 train_time_id=self.train_time_id,oot_time_id=self.oot_time_id,reject_time_id=self.reject_time_id,
                 record_list=self.record_list, model=self.model,scorecarddf=self.scorecarddf,model_var_type=self.model_var_type,
                 variable_list=self.variable_list,lasso_df=self.lasso_df,var_clus=self.var_clus)

    def plot_tab(self,
                 predict_train_data,predict_vaild_data,predict_oot_data,predict_reject_data,
                 train_target,oot_target,reject_target,
                 train_time_id,oot_time_id,reject_time_id,
                 record_list, model,scorecarddf,variable_list,lasso_df,model_var_type,var_clus):
        # root=self.master
        self.master.title('评分卡')#标题
        self.screenwidth = self.master.winfo_screenwidth()
        self.screenheight = self.master.winfo_screenheight()
        width=self.screenwidth*0.85
        height=self.screenheight*0.75
        print(width,height)
        alignstr = '%dx%d+%d+%d' % (width, height, (self.screenwidth -width ) /2, (self.screenheight -height ) /2)
        self.master.geometry(alignstr)

        self.master.resizable(width = True, height = True)#窗口大小

        def help1():
            try:
                self.master.destroy()
            except:
                pass

        def help2():
            # try:
            self.output_report()
            tkinter.messagebox.showinfo(title = '成功',message = '报告已成功道出到\n '+ self.project_path + '/' + '%s_model_report.xlsx' % self.node_name)
            # except Exception as e:
            #     tkinter.messagebox.showinfo(title='错误',message=e)

        menubar = Menu(self.master)
        filemenu = Menu(menubar, tearoff = 0)
        menubar.add_cascade(label = '菜单',menu = filemenu)
        filemenu.add_command(label='关闭',command = help1)
        filemenu.add_command(label='输出报告',command = help2)


        tabcontrol = ttk.Notebook(self.master)
        tab1 = ttk.Frame(tabcontrol)
        tabcontrol.add(tab1, text = '模型信息')
        tab2 = ttk.Frame(tabcontrol)
        tabcontrol.add(tab2, text = '模型表现')#Frame控件的具体用法
        tab3 = ttk.Frame(tabcontrol)
        tabcontrol.add(tab3, text = '稳定性表现')#Frame控件的具体用法
        tab4 = ttk.Frame(tabcontrol)
        tabcontrol.add(tab4, text = 'LASSO')#Frame控件的具体用法
        tab5 = ttk.Frame(tabcontrol)
        tabcontrol.add(tab5, text = '变量分组')#Frame控件的具体用法
        tab6 = ttk.Frame(tabcontrol)
        tabcontrol.add(tab6, text = '相关性分析')#Frame控件的具体用法
        tabcontrol.pack(expand=1, fill="both")

        #第一页模型信息   

        def processp(plist):
            x=''
            for m in plist:
                t=str(m)+'\n\n'
                x=x+t
            return x
#第一页模型表现
#   ----------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------
        process_bar_t=LabelFrame(tab1)
        scorecarddf['ave_score'] = scorecarddf.apply(lambda x: float(x['scorecard']) * float(x['pct_f_N_obs'].replace('%', '')) / 100,
                                   axis=1)
        av_df = scorecarddf.groupby('variable_name')['ave_score'].sum().reset_index().rename({'ave_score': 'average'}, axis=1)
        imp = pd.merge(scorecarddf, av_df, how='left', on='variable_name')
        imp['dif'] = abs(imp['scorecard'] - imp['average'])
        imp['dif_abs'] = imp.apply(lambda x: float(x['dif']) * float(x['pct_f_N_obs'].replace('%', '')) / 100, axis=1)
        show_df = imp.groupby('variable_name')['dif_abs'].sum().reset_index()
        show_df['im_pct'] = round(show_df['dif_abs'] / show_df['dif_abs'].sum(),2)
        def split_variable_name(x):
            i=0
            name=''
            while i< len(x):
                name+=x[i:i+20]
                name+='\n'
                i+=20
            return name
        show_df['name_split']=show_df['variable_name'].apply(lambda x: split_variable_name(x))
        show_df=show_df.sort_values(by='im_pct',ascending=False)
        try:
            plt.close()
        except:
            pass

        self.imp_gg = plt.figure(figsize=(min(max(20,len(show_df)*4),400),7),dpi=int((self.screenwidth/50)))
        plt.gcf().subplots_adjust(bottom=0.15)
        x = show_df['name_split']
        plt.bar(np.arange(len(x)) ,show_df['im_pct'],width=0.15)
        plt.xticks(np.arange(len(x)) , x, size=24)
        plt.tick_params(labelsize=24)
        plt.tight_layout()


        imp_var_base = LabelFrame(process_bar_t )
        imp_var_s = ScrollableFrame(imp_var_base)
        imp_var=LabelFrame(imp_var_s.scrollable_frame, text='变量重要性')
        imp_var_s.pack(side=BOTTOM, anchor='nw', fill=tk.BOTH, expand=YES)
        canvasg=FigureCanvasTkAgg(self.imp_gg,imp_var)
        canvasg.get_tk_widget().pack( expand=1)
        canvasg._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        imp_var.pack(side=TOP, fill='both',expand=1, anchor='ne')
        imp_var_base.pack(side=TOP, fill='both',expand=1, anchor='ne')

        process_bar=LabelFrame(process_bar_t,text='变量筛选过程')
        textPad1 = ScrolledText(process_bar,width =85,height =60)
        textPad1.insert(tkinter.constants.END, chars = processp(record_list) )
        textPad1.pack(side=tk.TOP,fill='both',expand=YES)
        process_bar.pack(side=TOP, fill='both',expand=NO, anchor='ne')
        process_bar_t.pack(side=RIGHT, fill='both',expand=YES, anchor='ne')
        var_list=max([len(x) for x in scorecarddf["variable_name"]])
        final=LabelFrame(tab1,text='最终模型')
        textPad = ScrolledText(final,width =var_list+80,height =30)
        textPad.insert(tkinter.constants.END, chars = str(model.summary()) )
        textPad.pack(fill='x',expand=NO)
        final.pack(side=TOP, anchor='w',fill='x',expand=1,  padx=5, pady=5)



        final2=LabelFrame(tab1,text='评分卡')
        columns = ('variable_name', 'f_group', 'label', 'f_Bad_rate', 'pct_f_N_obs',
                   'woe',
               'coff', 'scorecard')
        tree = ttk.Treeview(final2, show = "headings", 
                            columns = columns, selectmode = tk.BROWSE,height=70)
        tree.column("variable_name", anchor = "w",minwidth=0,width=150, stretch=NO)
        tree.column( 'f_group', anchor = "center",minwidth=0,width=50, stretch=NO)
        tree.column('label', anchor = "w",minwidth=0,width=200, stretch=NO)
        tree.column('f_Bad_rate', anchor = "center",minwidth=0,width=50, stretch=NO)
        tree.column('pct_f_N_obs', anchor = "center",minwidth=0,width=50, stretch=NO)
        tree.column('woe', anchor = "center",minwidth=0,width=50, stretch=NO)
        tree.column('coff', anchor = "center",minwidth=0,width=50, stretch=NO)
        tree.column('scorecard', anchor = "center",minwidth=0,width=50, stretch=NO)
        tree.heading("variable_name", text = "变量名称")
        tree.heading( 'f_group', text = "分组")
        tree.heading('label', text = "注释")
        tree.heading('f_Bad_rate', text = "坏账率")
        tree.heading('pct_f_N_obs', text = "样本占比")
        tree.heading('woe', text = "WOE")
        tree.heading('coff', text = "系数")
        tree.heading('scorecard', text = "分数")
        scorecarddf.apply(lambda x: tree.insert('',1, values = (x['variable_name'], x['f_group'], x['label'], x['f_Bad_rate'], x['pct_f_N_obs'],x['woe'],x['coff'], x['scorecard'])),axis=1)
        tree.pack(side=TOP,fill=tk.BOTH,expand=NO)
        final2.pack(side=TOP, anchor='w',fill='both',expand=NO,  padx=5, pady=5)

        preds_t=self.predict_train_data[self.train_target]
        labels_t=self.predict_train_data['SCORECARD_LR_p_1']
        #train------------------------------------
        if predict_vaild_data.empty==True:
            flag_v=False
            preds_v=()
            labels_v=()
        else:
            flag_v=True
            preds_v=self.predict_vaild_data[self.train_target]
            labels_v=self.predict_vaild_data['SCORECARD_LR_p_1']
         #reject  --------------------------------------------- 
        if (self.predict_reject_data.empty==True)|(self.reject_target==None):
            preds_r=()
            labels_r=()
            flag_reject_target=False
            flag_reject_data=False
        else:
            print(self.reject_target)
            preds_r=self.predict_reject_data[self.reject_target]
            labels_r=self.predict_reject_data['SCORECARD_LR_p_1']
            flag_reject_target=True
            flag_reject_data=True
        #oot------------------------------------------------------
        if (self.predict_oot_data.empty==True)|(self.oot_target==None):
            preds_o=()
            labels_o=()
            flag_oot_target=False
            flag_oot_data=False
        else:
            preds_o=self.predict_oot_data[self.oot_target]
            labels_o=self.predict_oot_data['SCORECARD_LR_p_1']
            flag_oot_target=True
            flag_oot_data=True
        print('a')
#第二页模型表现
# ----------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------
 #下1       -----------------------------------------------------------------------------
        middle_frame=Frame(tab2,width=self.screenwidth*0.72, height=self.screenheight*0.27)
        middle_frame.pack_propagate(0)
        ks=LabelFrame(middle_frame,text='KS')
            
        self.ks_mp=self.PlotKS(preds_t=preds_t, labels_t=labels_t, n=100, asc=0,preds_v=preds_v, labels_v=labels_v,
               preds_o=preds_o, labels_o=labels_o,preds_r=preds_r, labels_r=labels_r,
                flag_oot_data=flag_oot_data,flag_oot_target=flag_oot_target,flag_reject_data=flag_reject_data, flag_reject_target=flag_reject_target,flag_v=flag_v)
        canvasg=FigureCanvasTkAgg(self.ks_mp,ks)
        canvasg.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        canvasg._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        ks.pack(side=LEFT, anchor='nw',fill=tk.BOTH,expand=YES,  padx=5, pady=5)
    
        cali_data=LabelFrame(middle_frame,text='分数校准数据')
        if flag_v==True:
            columns = ('s_group','SCORECARD_LR_p_1',train_target,'count','SCORECARD_LR_p_1_v',train_target+'_v','count_v')
        else:
            columns = ('s_group','SCORECARD_LR_p_1',train_target,'count')
        tree = ttk.Treeview(cali_data, show = "headings", 
                            columns = columns, selectmode = tk.BROWSE)
        tree.column('s_group', anchor = "center",minwidth=0,width=80, stretch=NO)
        tree.column( 'SCORECARD_LR_p_1', anchor = "center",minwidth=0,width=60, stretch=NO)
        tree.column(train_target, anchor = "center",minwidth=0,width=60, stretch=NO)
        tree.column('count', anchor = "center",minwidth=0,width=60, stretch=NO)
        if flag_v==True:
            tree.column( 'SCORECARD_LR_p_1_v', anchor = "center",minwidth=0,width=60, stretch=NO)
            tree.column(train_target+'_v', anchor = "center",minwidth=0,width=60, stretch=NO)
            tree.column('count_v', anchor = "center",minwidth=0,width=60, stretch=NO)            
        tree.heading('s_group', text = "分数段")
        tree.heading( 'SCORECARD_LR_p_1', text = "Trian预测值")
        tree.heading(train_target, text = "Trian坏账率")
        tree.heading('count', text = "Trian样本数")
        if flag_v==True:
            tree.heading( 'SCORECARD_LR_p_1_v', text = "vaild预测值")
            tree.heading(train_target+'_v', text = "vaild预测值坏账率")
            tree.heading('count_v', text = "vaild预测值样本数")        
        cail_data_re = self.cali(df=predict_train_data,df_v=predict_vaild_data,score='SCORECARD_LR_p_1' ,target=train_target,flag_v=flag_v)
        self.cail_data=cail_data_re[0]
        self.cail_data['s_group']=self.cail_data.apply(lambda x: str(round(x['s_min'],4))+'-'+str(round(x['s_max'],4)) , axis=1)
        self.cail_data['SCORECARD_LR_p_1']=round(self.cail_data['SCORECARD_LR_p_1'],4)
        self.cail_data[train_target]=round(self.cail_data[train_target],4)
        if flag_v==True:
            self.cail_data['SCORECARD_LR_p_1_v']=round(self.cail_data['SCORECARD_LR_p_1_v'],4)
            self.cail_data[train_target+'_v']=round(self.cail_data[train_target+'_v'],4)
        self.cailbra_data=self.cail_data.sort_values(by='s_group',ascending = False)
        if flag_v==True:
            self.cailbra_data.apply(lambda x: tree.insert('',0, values = (x['s_group'], x['SCORECARD_LR_p_1'], x[train_target], x['count'],x['SCORECARD_LR_p_1_v'], x[train_target+'_v'], x['count_v'])),axis=1)
        else:
            self.cailbra_data.apply(lambda x: tree.insert('',0, values = (x['s_group'], x['SCORECARD_LR_p_1'], x[train_target], x['count'])),axis=1)
        tree.pack(fill='both',expand=YES)
        cali_data.pack(side=LEFT, anchor='nw',fill='y',expand=NO,  padx=5, pady=5)

        pltcali=LabelFrame(middle_frame,text='分数校准')
        # m = cali(df,'SCORECARD_LR_p_1' ,model1.target_train)
        self.cail_m=cail_data_re[1]
        canvasg=FigureCanvasTkAgg(self.cail_m,pltcali)
        canvasg.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        canvasg._tkcanvas.pack(side=RIGHT, fill=tk.BOTH, expand=1)
        pltcali.pack(side=RIGHT, anchor='w',fill='x',expand=YES,  padx=5, pady=5)


        middle_frame.pack(side=TOP, anchor='w',fill='x',expand=YES,  padx=5, pady=5)

 #下2       -----------------------------------------------------------------------------
        right_frame=Frame(tab2,width=self.screenwidth*0.72, height=self.screenheight*0.22)
        right_frame.pack_propagate(0)
        pltauc=LabelFrame(right_frame,text='AUC')


        self.auc_m=self.AUC(preds_t=preds_t, labels_t=labels_t, n=100, asc=0,preds_v=preds_v, labels_v=labels_v,
               preds_o=preds_o, labels_o=labels_o,preds_r=preds_r, labels_r=labels_r,
                flag_oot_data=flag_oot_data,flag_oot_target=flag_oot_target,flag_reject_data=flag_reject_data, flag_reject_target=flag_reject_target,flag_v=flag_v)
   
            
        canvasg=FigureCanvasTkAgg(self.auc_m,pltauc)
        canvasg.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        canvasg._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        pltauc.pack(side=LEFT, anchor='nw',fill=tk.BOTH,expand=YES,  padx=5, pady=5)
    
        lift_da=LabelFrame(right_frame,text='Lift数据')
        if flag_v==True:
            columns = ('total_pct','lift','total_pct_v','lift_v')
        else:
            columns = ('total_pct','lift')
        tree = ttk.Treeview(lift_da, show = "headings", columns = columns, selectmode = tk.BROWSE)
        tree.column('total_pct', anchor = "center",minwidth=0,width=80, stretch=NO)
        tree.column( 'lift', anchor = "center",minwidth=0,width=80, stretch=NO)
        if flag_v==True:
            tree.column('total_pct_v', anchor = "center",minwidth=0,width=80, stretch=NO)
            tree.column( 'lift_v', anchor = "center",minwidth=0,width=80, stretch=NO)            
        tree.heading('total_pct', text = "Train样本累计占比")
        tree.heading( 'lift', text = "Train提升率")
        if flag_v==True:
            tree.heading('total_pct_v', text = "Valid样本累计占比")
            tree.heading( 'lift_v', text = "Vliad提升率")
        temp2=self.cail_data.sort_values(by='s_group',ascending = False)
        temp2['total']=temp2['count'].cumsum(axis=0)
        temp2['bad']=round(temp2['count']*temp2[train_target],0)
        temp2['total_pct']=temp2['total']/temp2['count'].sum()
        temp2['lift']=temp2[train_target]/(temp2['bad'].sum()/temp2['count'].sum())
        temp2['total_pct']=round(temp2['total_pct'],2)
        temp2['lift']=round(temp2['lift'],2)
        if flag_v==True:
            temp2['total_v']=temp2['count_v'].cumsum(axis=0)
            temp2['bad_v']=round(temp2['count_v']*temp2[train_target+'_v'],0)
            temp2['total_pct_v']=temp2['total_v']/temp2['count_v'].sum()
            temp2['lift_v']=temp2[train_target+'_v']/(temp2['bad_v'].sum()/temp2['count_v'].sum())
            temp2['total_pct_v']=round(temp2['total_pct_v'],2)
            temp2['lift_v']=round(temp2['lift_v'],2)
        self.lift_data=temp2.sort_values(by='total_pct',ascending = False)
        if flag_v==True:   
            self.lift_data.apply(lambda x: tree.insert('',0, values = (x['total_pct'], x['lift'],x['total_pct_v'], x['lift_v'])),axis=1)
        else:
            self.lift_data.apply(lambda x: tree.insert('',0, values = (x['total_pct'], x['lift'])),axis=1)
            
        tree.pack(fill='both',expand=YES)
        lift_da.pack(side=LEFT, anchor='nw',fill='y',expand=NO,  padx=5, pady=5)



        pltlift=LabelFrame(right_frame,text='Lift')

        try:
            plt.close()
        except:
            pass
        self.lift_m=plt.figure()
        X=temp2['total_pct']
        Y=temp2['lift']#定义折线图的X，Y坐标
        plt.plot(X, Y,label='训练集') #折线图
        if flag_v==True:
            X_v=temp2['total_pct_v']
            Y_v=temp2['lift_v']#定义折线图的X，Y坐标
            plt.plot(X_v, Y_v,label='验证集') #折线图            
        i=0 
        for a, b in zip(X, Y):
            if (i % 10 == 0) |(i==5):
                plt.text(a, b,"%.0f%%" % (a * 100)+'--'+ '%.1f' %b , ha='center', va='bottom', fontsize=10)#每个点的数值
            i=i+1
        plt.legend()#显示每根折线的label
        plt.title('lift')#显示图名
        canvasg=FigureCanvasTkAgg(self.lift_m,pltlift)
        canvasg.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        canvasg._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        pltlift.pack(side=RIGHT, anchor='w',fill='x',expand=YES,  padx=5, pady=5)
        right_frame.pack(side=TOP, fill=tk.BOTH,expand=YES,  padx=5, pady=5)

        
 #下3       -----------------------------------------------------------------------------
        bottom_frame=Frame(tab2,width=self.screenwidth*0.72, height=self.screenheight*0.22)
        bottom_frame.pack_propagate(0)
        pltdistri=LabelFrame(bottom_frame,text='分数分布')
        self.dis_train_m = self.plotdis(predict_train_data,train_target,'SCORECARD_LR_p_1')
        canvasg=FigureCanvasTkAgg(self.dis_train_m,pltdistri)
        canvasg.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        canvasg._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        pltdistri.pack(side=LEFT, anchor='nw',fill='y',expand=NO,  padx=5, pady=5)

    
        disgb=LabelFrame(bottom_frame,text='分数区间分布')
        

        if self.predict_oot_data.empty==True:
            oot_data=False
        else:
            oot_data=True
        if self.oot_target==None:
            oot_target_flag=False
        else:
            oot_target_flag=True
            
        if self.predict_reject_data.empty==True:
            reject_data=False
        else:
            reject_data=True
        if self.reject_target==None:
            reject_target_flag=False
        else:
            reject_target_flag=True
            
        if self.predict_vaild_data.empty==True:
            vaild_data=False
        else:
            vaild_data=True
        self.all_sample_dis_m= self.disallfb(df=self.predict_train_data,df_v=self.predict_vaild_data,df_t=self.predict_oot_data,df_r=self.predict_reject_data,
                                 oot_data=oot_data,oot_target_flag=oot_target_flag,vaild_data=vaild_data,reject_data=reject_data,reject_target_flag=reject_target_flag,
                                 train_target=self.train_target,reject_target=self.reject_target,oot_target=self.oot_target)
        canvasg=FigureCanvasTkAgg(self.all_sample_dis_m,disgb)
        canvasg.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        canvasg._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        disgb.pack(side=TOP, anchor='nw',fill=tk.BOTH,expand=YES,  padx=5, pady=5)
        bottom_frame.pack(side=TOP, fill=tk.BOTH,expand=YES,  padx=5, pady=5)
        print('b')
#第三页模型表现
#   ----------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------

        if (train_time_id!=None ) or (reject_time_id!=None and predict_reject_data.empty==False) or( oot_time_id!=None and predict_oot_data.empty==False) :
            left_frame_3=Frame(tab3,width=self.screenwidth*0.32, height=self.screenheight*0.27)
            self.left_frame_3_width=self.screenwidth*0.32
            self.model_ginitime=self.calculateginitime('SCORECARD_LR_p_1', predict_train_data, train_time_id, train_target, flag_v,predict_vaild_data,
                                    predict_oot_data, oot_target, oot_time_id, predict_reject_data, reject_time_id,reject_target)

    #-------------------------------------------------------------------------------------------
            if self.model_ginitime.empty == False:
                steab_data=LabelFrame(left_frame_3,text='模型表现稳定性')
                col=[]
                if train_time_id != None:
                    col=['timeid','auc','count','bad_rate']
                if (flag_v==True) and (train_time_id!= None):
                    col=col+['auc_v','count_v','bad_rate_v']
                if self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None :
                    col=col+['auc_oot','count_oot','bad_rate_oot']
                if self.predict_reject_data.empty==False and self.reject_target!=None and self.reject_time_id!=None :
                    col=col+['auc_reject','count_reject','bad_rate_reject']
                columns = col
                tree = ttk.Treeview(steab_data, show = "headings",
                                    columns = columns, selectmode = tk.BROWSE)
                tree.column('timeid', anchor = "center",minwidth=0,width=80, stretch=NO)
                if train_time_id != None:
                    tree.column( 'auc', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree.column('count', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree.column('bad_rate', anchor = "center",minwidth=0,width=60, stretch=NO)
                if (flag_v==True) and (train_time_id!= None):
                    tree.column( 'auc_v', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree.column('count_v', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree.column('bad_rate_v', anchor = "center",minwidth=0,width=60, stretch=NO)
                if self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None :
                    tree.column( 'auc_oot', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree.column('count_oot', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree.column('bad_rate_oot', anchor = "center",minwidth=0,width=60, stretch=NO)
                if self.predict_reject_data.empty==False and self.reject_target!=None and self.reject_time_id!=None :
                    tree.column( 'auc_reject', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree.column('count_reject', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree.column('bad_rate_reject', anchor = "center",minwidth=0,width=60, stretch=NO)
                tree.heading('timeid', text = "时间ID")
                if train_time_id != None:
                    tree.heading( 'auc', text = "Trian AUC")
                    tree.heading('count', text = "Trian样本数")
                    tree.heading('bad_rate', text = "Trian坏账率")
                if (flag_v==True) and (train_time_id!= None):
                    tree.heading('auc_v', text = "vaild AUC")
                    tree.heading('count_v', text = "vaild样本数")
                    tree.heading('bad_rate_v', text = "vaild坏账率")
                if self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None :
                    tree.heading('auc_oot', text = "OOT AUC")
                    tree.heading('count_oot', text = "OOT样本数")
                    tree.heading('bad_rate_oot', text = "OOT坏账率")
                if self.predict_reject_data.empty==False and self.reject_target!=None and self.reject_time_id!=None :
                    tree.heading('auc_reject', text = "Reject AUC")
                    tree.heading('count_reject', text = "Reject样本数")
                    tree.heading('bad_rate_reject', text = "Reject坏账率")
                def inter(x):
                    re=(x['timeid'],)
                    if train_time_id != None:
                        re=re+(x['auc'],x['count'],x['bad_rate'],)
                    if (flag_v==True) and (train_time_id!= None):
                        re=re+(x['auc_v'],x['count_v'],x['bad_rate_v'],)
                    if self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None:
                        re=re+(x['auc_oot'],x['count_oot'],x['bad_rate_oot'],)
                    if self.predict_reject_data.empty==False and self.reject_target!=None and self.reject_time_id!=None :
                        re=re+(x['auc_reject'],x['count_reject'],x['bad_rate_reject'],)
                    return re
                self.model_ginitime.apply(lambda x: tree.insert('',0, values = inter(x)),axis=1)
                tree.pack(fill='both',expand=YES)
                steab_data.pack(side=TOP, anchor='nw',fill='y',expand=NO,  padx=5, pady=5)
        #-------------------------------------------------------------------------------------------
                steab_fig=LabelFrame(left_frame_3,text='模型表现稳定性')
                try:
                    plt.close()
                except:
                    pass
                self.gg_time = plt.figure(figsize=(10,3),dpi=int((self.left_frame_3_width/10)))
                ax1 = self.gg_time.add_subplot(1, 1, 1)
                plt.subplots_adjust(left=None, bottom=None, right=None, top=None,
                                    wspace=1, hspace=None)
                self.model_ginitime = self.model_ginitime.sort_values(by='timeid', ascending=True)
                x = self.model_ginitime['timeid']

                bar_width = 0.1
                if train_time_id != None:
                    y1 = self.model_ginitime['bad_rate']
                    plt.bar(x=range(len(x)), height=y1, label='Train',  width=bar_width)
                if (flag_v==True) and (train_time_id!= None):
                    y2 = self.model_ginitime['bad_rate_v']
                    plt.bar(x=np.arange(len(x)) + bar_width, height=y2, label='valid',
                            width=bar_width)
                if self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None :
                    y3 = self.model_ginitime['bad_rate_oot']
                    plt.bar(x=np.arange(len(x)) + flag_v * bar_width+bar_width, height=y3, label='reject',
                            width=bar_width)
                if self.predict_reject_data.empty==False and self.reject_target!=None and self.reject_time_id!=None :
                    flag_oot=self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None
                    y4 = self.model_ginitime['bad_rate_reject']
                    plt.bar(x=np.arange(len(x)) + flag_v * bar_width+flag_oot*bar_width+bar_width, height=y4, label='oot',
                            width=bar_width)

                plt.legend()
                # 为两条坐标轴设置名称
                plt.xlabel(u"时间",fontsize=12)
                plt.ylabel(u"坏账率",fontsize=12)

                ax2 = ax1.twinx()
                if train_time_id != None:
                    auc_t = self.model_ginitime['auc']
                    ax2.plot(np.arange(len(x)), auc_t)
                if (flag_v==True) and (train_time_id!= None):
                    auc2 =  self.model_ginitime['auc_v']
                    ax2.plot(np.arange(len(x)) + bar_width, auc2)
                if self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None:
                    badrate3 = self.model_ginitime['auc_oot']
                    ax2.plot(np.arange(len(x)) + bar_width *flag_v+bar_width, badrate3 )
                if self.predict_reject_data.empty==False and self.reject_target!=None and self.reject_time_id!=None :
                    badrate4 = self.model_ginitime['auc_reject']
                    flag_oot=self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None
                    ax2.plot(np.arange(len(x)) + bar_width *flag_v+bar_width*flag_oot+bar_width, badrate4 )
                ax2.set_ylabel('AUC',fontsize=12)
                plt.xticks(np.arange(len(x)) + 0.15, x, fontsize=12, rotation=45)
                plt.tight_layout()
                canvasg=FigureCanvasTkAgg(self.gg_time,steab_fig)
                canvasg.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=0)
                canvasg._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=0)
                steab_fig.pack(side=TOP, anchor='nw',fill=tk.BOTH,expand=0,  padx=5, pady=5)
            steab_dis = LabelFrame(left_frame_3, text='分布稳定性')
            if self.predict_oot_data.empty == True:
                oot_data = False
            else:
                oot_data = True
            if self.predict_reject_data.empty == True:
                reject_data = False
            else:
                reject_data = True

            if self.predict_vaild_data.empty == True:
                vaild_data = False
            else:
                vaild_data = True
            self.gg_dis=self.distimeall(df=self.predict_train_data, df_v=self.predict_vaild_data, df_t=self.predict_oot_data, df_r=self.predict_reject_data,
                       oot_data=oot_data,  vaild_data=vaild_data, reject_data=reject_data, score='SCORECARD_LR_p_1', train_time_id=self.train_time_id, oot_time_id=self.oot_time_id, reject_time_id=self.reject_time_id)
            canvasg=FigureCanvasTkAgg(self.gg_dis,steab_dis)
            canvasg.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=0)
            canvasg._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=0)
            steab_dis.pack(side=TOP, anchor='nw',fill=tk.BOTH,expand=0,  padx=5, pady=5)
            left_frame_3.pack(side=LEFT, anchor='nw',fill=tk.BOTH,expand=0,  padx=5, pady=5)

    #第三页右边----------------------------------------------------------------------------------------------------

            right_frame_b3=Frame(tab3,width=self.screenwidth*0.5, height=self.screenheight*0.27)
            self.right_frame_b3_width=self.screenwidth*0.4

            def plot_variable_sta(event):
                try:
                    self.s_frame.pack_forget()
                    self.s_frame = ScrollableFrame(right_frame_b3)
                except:
                    self.s_frame = ScrollableFrame(right_frame_b3)
                var_in=self.comboxlist_variable_use.get()
                if var_in !='const':
                    if list(scorecarddf[scorecarddf['variable_name']==var_in]['var_type'])[0]=='add':
                        if model_var_type=='WOE':
                            # model_ginitime = calculateginitime(var_in)
                            model_ginitime = self.calculateginitime(var_in, predict_train_data,
                                                                         train_time_id, train_target, flag_v,
                                                                         predict_vaild_data,
                                                                         predict_oot_data, oot_target, oot_time_id,
                                                                         predict_reject_data, reject_time_id,
                                                                         reject_target)
                        else:
                            # model_ginitime=calculateginitime('f_group_%s'%var_in)
                            model_ginitime = self.calculateginitime('f_group_%s'%var_in, predict_train_data,
                                                                         train_time_id, train_target, flag_v,
                                                                         predict_vaild_data,
                                                                         predict_oot_data, oot_target, oot_time_id,
                                                                         predict_reject_data, reject_time_id,
                                                                         reject_target)
                    else:
                        # model_ginitime=calculateginitime('woe_%s'%var_in)
                        model_ginitime = self.calculateginitime('woe_%s'%var_in, predict_train_data,
                                                                     train_time_id, train_target, flag_v,
                                                                     predict_vaild_data,
                                                                     predict_oot_data, oot_target, oot_time_id,
                                                                     predict_reject_data, reject_time_id, reject_target)
                    var_in_frame=LabelFrame(self.s_frame.scrollable_frame,text=var_in)
                    col=['timeid']
                    if train_time_id != None:
                        col=col+['auc','count','bad_rate']
                    if (flag_v==True) and (train_time_id!= None):
                        col=col+['auc_v','count_v','bad_rate_v']
                    if self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None :
                        col=col+['auc_oot','count_oot','bad_rate_oot']
                    if self.predict_reject_data.empty==False and self.reject_target!=None and self.reject_time_id!=None :
                        col=col+['auc_reject','count_reject','bad_rate_reject']
                    columns = col
                    tree = ttk.Treeview(var_in_frame, show = "headings",
                                        columns = columns, selectmode = tk.BROWSE)
                    tree.column('timeid', anchor = "center",minwidth=0,width=80, stretch=NO)
                    if train_time_id != None:
                        tree.column( 'auc', anchor = "center",minwidth=0,width=60, stretch=NO)
                        tree.column('count', anchor = "center",minwidth=0,width=60, stretch=NO)
                        tree.column('bad_rate', anchor = "center",minwidth=0,width=60, stretch=NO)
                    if (flag_v==True) and (train_time_id!= None):
                        tree.column( 'auc_v', anchor = "center",minwidth=0,width=60, stretch=NO)
                        tree.column('count_v', anchor = "center",minwidth=0,width=60, stretch=NO)
                        tree.column('bad_rate_v', anchor = "center",minwidth=0,width=60, stretch=NO)
                    if self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None :
                        tree.column( 'auc_oot', anchor = "center",minwidth=0,width=60, stretch=NO)
                        tree.column('count_oot', anchor = "center",minwidth=0,width=60, stretch=NO)
                        tree.column('bad_rate_oot', anchor = "center",minwidth=0,width=60, stretch=NO)
                    if self.predict_reject_data.empty==False and self.reject_target!=None and self.reject_time_id!=None :
                        tree.column( 'auc_reject', anchor = "center",minwidth=0,width=60, stretch=NO)
                        tree.column('count_reject', anchor = "center",minwidth=0,width=60, stretch=NO)
                        tree.column('bad_rate_reject', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree.heading('timeid', text = "时间ID")
                    if train_time_id != None:
                        tree.heading( 'auc', text = "Trian AUC")
                        tree.heading('count', text = "Trian样本数")
                        tree.heading('bad_rate', text = "Trian坏账率")
                    if (flag_v==True) and (train_time_id!= None):
                        tree.heading('auc_v', text = "vaild AUC")
                        tree.heading('count_v', text = "vaild样本数")
                        tree.heading('bad_rate_v', text = "vaild坏账率")
                    if self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None :
                        tree.heading('auc_oot', text = "OOT AUC")
                        tree.heading('count_oot', text = "OOT样本数")
                        tree.heading('bad_rate_oot', text = "OOT坏账率")
                    if self.predict_reject_data.empty==False and self.reject_target!=None and self.reject_time_id!=None :
                        tree.heading('auc_reject', text = "Reject AUC")
                        tree.heading('count_reject', text = "Reject样本数")
                        tree.heading('bad_rate_reject', text = "Reject坏账率")
                    def inter(x):
                        re=(x['timeid'],)
                        if train_time_id != None:
                            re=re+(x['auc'],x['count'],x['bad_rate'],)
                        if (flag_v==True) and (train_time_id!= None):
                            re=re+(x['auc_v'],x['count_v'],x['bad_rate_v'],)
                        if self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None:
                            re=re+(x['auc_oot'],x['count_oot'],x['bad_rate_oot'],)
                        if self.predict_reject_data.empty==False and self.reject_target!=None and self.reject_time_id!=None :
                            re=re+(x['auc_reject'],x['count_reject'],x['bad_rate_reject'],)
                        return re
                    model_ginitime.apply(lambda x: tree.insert('',0, values = inter(x)),axis=1)
                    tree.pack(fill='both',expand=YES,  padx=5, pady=5)
    #报告----------------------------------------------------------------------------------------
    #--------------------------------------------------------------------------------------
                    if list(scorecarddf[scorecarddf['variable_name']==var_in]['var_type'])[0]=='add':
                        report=scorecarddf[scorecarddf['variable_name']==var_in][ ['f_group','label', 'f_Bad_rate', 'f_N_obs']]
                        report['iv']=None
                        report['miss_rate']=0
                    else:
                        report=self.f_group_report[self.f_group_report['variable_name']==var_in]
                    columns=[ 'f_group','label','iv', 'f_Bad_rate', 'f_N_obs','miss_rate']
                    tree_rep = ttk.Treeview(var_in_frame, show = "headings",
                                        columns = columns, selectmode = tk.BROWSE)
                    tree_rep.column( 'f_group', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree_rep.column('label', anchor = "center",minwidth=0,width=280, stretch=NO)
                    tree_rep.column('iv', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree_rep.column('f_Bad_rate', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree_rep.column('f_N_obs', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree_rep.column('miss_rate', anchor = "center",minwidth=0,width=60, stretch=NO)
                    tree_rep.heading( 'f_group', text = "组数")
                    tree_rep.heading('label', text = "标识")
                    tree_rep.heading('iv', text = "信息熵")
                    tree_rep.heading('f_Bad_rate', text = "组坏账率")
                    tree_rep.heading('f_N_obs', text = "组坏账数")
                    tree_rep.heading('miss_rate', text = "缺失率")
                    report.apply(lambda x: tree_rep.insert('',0, values =(x['f_group'],x['label'],x['iv'], x['f_Bad_rate'], x['f_N_obs'],x['miss_rate'])),axis=1)
                    tree_rep.pack(fill='both',expand=YES,  padx=5, pady=5)

            #woe------group--------
                    ccc=self.groupplot_datainit(modify_var=var_in)
                    self.ccc=ccc
                    re_fig=self.plot_group(bad_pct=ccc[0], total_pct=ccc[1], final=ccc[2])
                    canvasg=FigureCanvasTkAgg(re_fig,var_in_frame)
                    canvasg.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
                    canvasg._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
                    var_in_frame.pack(side=TOP, anchor='nw',fill=tk.BOTH,expand=YES,  padx=5, pady=5)
                    self.s_frame.pack(side=BOTTOM, anchor='nw',fill=tk.BOTH,expand=YES)
            var_list=list(set(scorecarddf['variable_name']))
            try:
                var_list.remove('const')
            except:
                pass
            toplbel=LabelFrame(right_frame_b3)
            self.comboxlist_variable_use = ttk.Combobox(toplbel, width=45)
            self.comboxlist_variable_use["value"] = var_list
    #         self.comboxlist_variable_use.current(0)
            self.comboxlist_variable_use.bind("<<ComboboxSelected>>",plot_variable_sta )
            self.comboxlist_variable_use.pack(side=RIGHT, fill='x', anchor='ne', padx=3, pady=3)
            labelc = ttk.Label(toplbel, text='变量选择：')
            labelc.pack(side=RIGHT, fill='x', anchor='ne', padx=3, pady=3)
            toplbel.pack(side=TOP, fill='x', anchor='ne', padx=3, pady=3)
            right_frame_b3.pack(side=RIGHT, anchor='nw',fill=tk.BOTH,expand=YES,  padx=5, pady=5)
        print('c')
 #第四页-------------------------------------------------------------------------
#-----------------------------------------------------------------------------------
        if lasso_df.empty==False:
            lasso_bar_b=LabelFrame(tab4)
            self.s_frame_lasso = ScrollableFrame(lasso_bar_b)
            lasso_bar=LabelFrame(self.s_frame_lasso.scrollable_frame,text='LASSO')
            woe_list=['woe_'+ x for x in variable_list]
            summ=lasso_df.sort_values(by='C')
            summ=summ.reset_index().drop(columns='index')
            remove_list=[]
            var_list=woe_list
            for var in var_list:
                if summ[var].sum()==0:
                    remove_list.append(var)
            setp_df=summ.groupby('k')['aic'].max().reset_index().reset_index().rename({'index':'step'},axis=1)
            final=pd.merge(setp_df,summ,how='left',on=['k','aic'])
            final=final.fillna(0)
            self.final_df=final.copy()
            for step_va in range(final['step'].min(),final['step'].max()):
                if list(final.loc[final['step']==step_va]['aic'])[0]<final[(final['step']>step_va)&(final['step']<=step_va+5)]['aic'].min():
                    stop_point_aic=step_va
                    break
                else:
                    stop_point_aic=step_va
            for step_va in range(final['step'].min(),final['step'].max()):
                if list(final.loc[final['step']==step_va]['bic'])[0]<final[(final['step']>step_va)&(final['step']<=step_va+5)]['bic'].min():
                    stop_point_bic=step_va
                    break
                else:
                    stop_point_aic=step_va
            var_list=set(final.columns)-set(['llr', 'aic', 'bic','k','index',  'C','step','llr_v','aic_v','bic_v'])-set(remove_list)
            step_show_list=[]
            for vari_t in var_list:
                min_setp=final[final[vari_t]!=0]['step'].min()
                dic_df={'variable_name':vari_t,'step':min_setp}
                step_show_list.append(dic_df)
            step_show_df=pd.DataFrame(step_show_list)
            if ('aic_v' in final.columns) and( 'bic_v' in final.columns):
                df_final_mr=final[['step','llr', 'aic', 'bic','llr_v','aic_v','bic_v']]
            else:
                df_final_mr=final[['step','llr', 'aic', 'bic']]
            step_show_df=pd.merge(step_show_df,df_final_mr,how='left',on='step')
            step_show_df=step_show_df.sort_values(by='step',ascending=False)
            if ('aic_v' in final.columns) and( 'bic_v' in final.columns):
                columns=[ 'variable_name','step','llr', 'aic', 'bic','llr_v','aic_v','bic_v']
            else:
                columns=[ 'variable_name','step','llr', 'aic', 'bic']
            lasso_df=LabelFrame(lasso_bar,text='LASSO 数据')
            tree_lasso = ttk.Treeview(lasso_df, show = "headings",
                                columns = columns, selectmode = tk.BROWSE)
            tree_lasso.column( 'variable_name', anchor = "center",minwidth=0,width=320, stretch=NO)
            tree_lasso.column('step', anchor = "center",minwidth=0,width=50, stretch=NO)
            tree_lasso.column('llr', anchor = "center",minwidth=0,width=50, stretch=NO)
            tree_lasso.column('bic', anchor = "center",minwidth=0,width=50, stretch=NO)
            tree_lasso.column('aic', anchor = "center",minwidth=0,width=50, stretch=NO)
            if ('aic_v' in final.columns) and( 'bic_v' in final.columns):
                tree_lasso.column('llr_v', anchor = "center",minwidth=0,width=50, stretch=NO)
                tree_lasso.column('bic_v', anchor = "center",minwidth=0,width=50, stretch=NO)
                tree_lasso.column('aic_v', anchor = "center",minwidth=0,width=50, stretch=NO)
            tree_lasso.heading( 'variable_name', text = "变量名")
            tree_lasso.heading('step', text = "步数")
            tree_lasso.heading('llr', text = "训练集LLR")
            tree_lasso.heading('aic', text = "训练集AIC")
            tree_lasso.heading('bic', text = "训练集BIC")
            if ('aic_v' in final.columns) and( 'bic_v' in final.columns):
                tree_lasso.heading('llr_v', text = "验证集LLR")
                tree_lasso.heading('aic_v', text = "验证集AIC")
                tree_lasso.heading('bic_v', text = "验证集BIC")
            self.step_show_df=step_show_df.copy()

            if ('aic_v' in final.columns) and( 'bic_v' in final.columns):
                step_show_df.apply(lambda x: tree_lasso.insert('',0, values =(x['variable_name'],int(x['step']),int(x['llr']), int(x['aic']),
                                                                              int(x['bic']),int(x['llr_v']),int(x['aic_v']),int(x['bic_v']))),axis=1)
            else:
                step_show_df.apply(lambda x: tree_lasso.insert('',0, values =(x['variable_name'],int(x['step']),int(x['llr']), int(x['aic']),
                                                                              int(x['bic']))),axis=1)
            tree_lasso.pack(side=TOP,fill=tk.BOTH,expand=YES,  padx=5, pady=5)
            try:
                plt.close()
            except:
                pass
            self.lasso_fig = plt.figure(figsize=(8,4*3),dpi=int((self.screenwidth/10)))
            final=final.fillna(0)
            ax1=self.lasso_fig.add_subplot(3, 1,1)
            plt.subplots_adjust(left=None, bottom=None, right=None, top=None,
                        wspace=0.3, hspace=0.2)
            box = ax1.get_position()
            ax1.set_position([box.x0, box.y0, box.width , box.height* 0.8])
            plt.xlabel('步数')
            x = final['step']
            y1 = final[['aic']]
            y2 = final[['bic']]
            y3 = final[['llr']]
            plt.ylabel('系数')
            for varia in var_list:
                plt.plot(x, final[varia], label=varia)
            plt.axvline(stop_point_aic, color='gray', linestyle='--',label='aic_stop_point')
            plt.axvline(stop_point_bic, color='orange', linestyle='--',label='bic_stop_point')

            self.lasso_fig.add_subplot(3, 1,2)

            if ('aic_v' in final.columns) and( 'bic_v' in final.columns):
                y1_v=final[['aic_v']]
                y2_v=final[['bic_v']]
                plt.plot(x, y1_v,color='blue',label='aic_vaild')
                plt.plot(x, y2_v,color='green',label='bic_vaild')
            plt.plot(x, y1,color='red',label='aic')
            plt.plot(x, y2,color='black',label='bic')
            plt.xlabel('步数')
            plt.legend()
            ax8=plt.twinx()
            ax8.plot(x, y3,color='blue',label='llr')
            plt.axvline(stop_point_aic, color='gray', linestyle='--',label='aic_stop_point')
            plt.axvline(stop_point_bic, color='orange', linestyle='--',label='bic_stop_point')
            plt.grid(axis="y")
            plt.legend()
            self.lasso_fig.add_subplot(3, 1,3)

            plt.xlabel('step')
            x = final[final['step']<=max(stop_point_aic,stop_point_bic)]['step']
            remove_list_b=[]
            for var in var_list:
                if final[final['step']<=max(stop_point_aic,stop_point_bic)][var].sum()==0:
                    remove_list_b.append(var)
            plt.ylabel('系数')
            var_list_b=var_list-set(remove_list_b)
            for varia in var_list_b:
                plt.plot(x, final[final['step']<=max(stop_point_aic,stop_point_bic)][varia], label=varia)

            plt.legend(bbox_to_anchor=(0, -0.1), loc="upper left", prop = {'size':6},
                     borderaxespad=0, ncol=3)
            plt.tight_layout()
            lasso_pig=LabelFrame(lasso_bar,text='LASSO 图表')
            canvasg=FigureCanvasTkAgg(self.lasso_fig,lasso_pig)
            canvasg.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
            canvasg._tkcanvas.pack(side=TOP, fill=tk.BOTH, expand=1)
            lasso_df.pack(side=TOP, fill=tk.BOTH, expand=1)
            lasso_pig.pack(side=TOP, fill=tk.BOTH, expand=1)
            lasso_bar.pack(side=TOP, fill=tk.BOTH,expand=YES)
            self.s_frame_lasso.pack(side=TOP, fill=tk.BOTH,expand=YES)
            lasso_bar_b.pack(side=TOP, fill=tk.BOTH,expand=YES,  padx=5, pady=5)
        print('d')
  #第五页-------------------------------------------------------------------------
#-----------------------------------------------------------------------------------       
        
        
        clus_bar=LabelFrame(tab5,text='变量分组')
        # variable_list=['woe_'+ x for x in variable_list]
        # df=predict_train_data[variable_list]
        # clus = VarClus()
        # clus.decompose(dataframe=df)
        # model_list=['woe_'+ x for x in list(set(scorecarddf['variable_name']))]
        # h=clus.print_cluster_structure(model_variable=model_list, h_space=5)
        textPad2 = ScrolledText(clus_bar,width =85,height =100)
        textPad2.insert(tkinter.constants.END, chars = var_clus )
        textPad2.pack(fill='both',expand=YES)
        clus_bar.pack(side=TOP, fill='both',expand=NO, anchor='ne')
        # 第六页-------------------------------------------------------------------------
        # -----------------------------------------------------------------------------------
        corr_tab = LabelFrame(tab6)
        mm = self.predict_train_data
        mm1 = self.scorecarddf
        variable = list(mm1['variable_name'].unique())

        woe_variable = set(['woe_' + x for x in variable])
        woe_variable = list(woe_variable.intersection(set(self.predict_train_data.columns)))
        def split_variable_name(x):
            i=0
            name=''
            while i< len(x):
                name+=x[i:i+15]
                name+='\n'
                i+=15
            return name
        sw_varibale=[split_variable_name(x) for x in woe_variable]
        self.corr_data = mm[woe_variable].corr()
        try:
            plt.close()
        except:
            pass

        g = plt.figure(figsize=(min(max(46,len(sw_varibale)*4),400),min(max(self.screenheight/self.screenwidth*44,len(sw_varibale)*4),400)),dpi=int(self.screenwidth/50))
        sns.heatmap(mm[woe_variable].corr(), linewidths=0.1, vmax=1.0,vmin=-1,
                    square=True, linecolor='white', annot=True,annot_kws={'size':24,'weight':'bold', 'color':'blue'})

        plt.xticks(np.arange(len(woe_variable))+0.25, sw_varibale, size=24,rotation=360)
        plt.yticks(np.arange(len(woe_variable))+0.25, sw_varibale, size=24)
        plt.tick_params(labelsize=24)


        corr_tab_base = LabelFrame(corr_tab )
        corr_tab_s = ScrollableFrame(corr_tab_base)
        corr_tab_g=LabelFrame(corr_tab_s.scrollable_frame,text='相关性分析')
        corr_tab_s.pack(side=BOTTOM, anchor='nw', fill=tk.BOTH, expand=YES)
        canvasg=FigureCanvasTkAgg(g,corr_tab_g)
        canvasg.get_tk_widget().pack( expand=1)
        canvasg._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        corr_tab_g.pack(side=TOP, fill='both',expand=1, anchor='ne')
        corr_tab_base.pack(side=TOP, fill='both',expand=1, anchor='ne')
        corr_tab.pack(side=TOP, anchor='nw', fill=tk.BOTH, expand=1, padx=5, pady=5)

        self.master.config(menu = menubar)
    def PlotKS(self,preds_t, labels_t, n, asc,preds_v, labels_v,
                   preds_o, labels_o,preds_r, labels_r,flag_oot_data,flag_oot_target,flag_reject_data, flag_reject_target,flag_v):

        def calculate_data(preds, labels):
            data=pd.DataFrame({'pred':preds,'y_label':labels})
            # data['label']=data['y_label'].apply(lambda x: 'Y' if x==1 else 'N' )
            crossfreq = pd.crosstab(data['pred'], data['y_label'])
            crossfreq['total'] = crossfreq[0] + crossfreq[1]
            crossdens = crossfreq.cumsum(axis=0) / crossfreq.sum()
            crossdens['gap'] = abs(crossdens[0] - crossdens[1])
            ks = crossdens[crossdens['gap'] == crossdens['gap'].max()]
            return ks , crossdens
        ks_t , crossdens_t=calculate_data( labels_t,preds_t)
        if flag_v==True:
            ks_v , crossdens_v=calculate_data( labels_v,preds_v)

        if flag_oot_data==True and flag_oot_target==True:
            ks_o , crossdens_o=calculate_data( labels_o,preds_o)
        if flag_reject_data==True and flag_reject_target==True:
            ks_r , crossdens_r=calculate_data( labels_r,preds_r)
        try:
            plt.close()
        except:
            pass
        gg=plt.figure()
        # chart
        plt.plot(crossdens_t['total'], crossdens_t[0], label='训练累计好占比',
                 color='blue', linestyle='-', linewidth=2)

        plt.plot(crossdens_t['total'], crossdens_t[1], label='训练累计坏占比',
                 color='red', linestyle='-', linewidth=2)

        plt.plot(crossdens_t['total'], crossdens_t['gap'], label='训练ks曲线',
                 color='green', linestyle='-', linewidth=2)
        if flag_v==True:
            plt.plot(crossdens_v['total'], crossdens_v[0], label='验证累计好占比',
             color='blue', linestyle='--', linewidth=2)

            plt.plot(crossdens_v['total'], crossdens_v[1], label='验证累计坏占比',
                     color='red', linestyle='--', linewidth=2)

            plt.plot(crossdens_v['total'], crossdens_v['gap'], label='验证ks曲线',
                     color='green', linestyle='--', linewidth=2)

        if flag_oot_data==True and flag_oot_target==True:
            plt.plot(crossdens_o['total'], crossdens_o[0], label='OOT累计好占比',
              linestyle='--', linewidth=2)

            plt.plot(crossdens_o['total'], crossdens_o[1], label='OOT累计坏占比',
                      linestyle='--', linewidth=2)

            plt.plot(crossdens_o['total'], crossdens_o['gap'], label='OOT ks曲线',
                      linestyle='--', linewidth=2)
        if flag_reject_data==True and flag_reject_target==True:
            plt.plot(crossdens_r['total'], crossdens_r[0], label='拒绝累计好占比',
              linestyle='--', linewidth=2)

            plt.plot(crossdens_r['total'], crossdens_r[1], label='拒绝累计坏占比',
                      linestyle='--', linewidth=2)

            plt.plot(crossdens_r['total'], crossdens_r['gap'], label='拒绝 ks曲线',
                      linestyle='--', linewidth=2)
        plt.legend()
        plt.axvline(list(ks_t['total'])[0], color='gray', linestyle='--')
        tiltle='训练KS=%s ' % np.round(list(ks_t['gap'])[0], 4) +'at Pop=%s' % np.round(list(ks_t['total'])[0], 4)
        if flag_v==True:
            tiltle=tiltle+'\n验证KS=%s ' % np.round(list(ks_v['gap'])[0], 4) +'at Pop=%s' % np.round(list(ks_v['total'])[0], 4)
        if flag_reject_data==True and flag_reject_target==True:
            tiltle=tiltle+'\n拒绝KS=%s ' % np.round(list(ks_r['gap'])[0], 4) +'at Pop=%s' % np.round(list(ks_r['total'])[0], 4)
        if flag_oot_data==True and flag_oot_target==True:
            tiltle=tiltle+'\nOOTKS=%s ' % np.round(list(ks_o['gap'])[0], 4) +'at Pop=%s' % np.round(list(ks_o['total'])[0], 4)
        plt.title(tiltle, fontsize=12)
        plt.tight_layout()
        return gg
    def AUC(self,preds_t, labels_t, n, asc,preds_v, labels_v,
                   preds_o, labels_o,preds_r, labels_r,flag_oot_data,flag_oot_target,flag_reject_data, flag_reject_target,flag_v):
        try:
            plt.close()
        except:
            pass
        fpr_t, tpr_t, threshold_t = roc_curve(preds_t,labels_t)  ###计算真正率和假正率
        roc_auc_t = auc(fpr_t, tpr_t)  ###计算auc的值
        if flag_v==True:
            fpr_v, tpr_v, threshold_v = roc_curve(preds_v,labels_v)  ###计算真正率和假正率
            roc_auc_v = auc(fpr_v, tpr_v)  ###计算auc的值   
        if flag_reject_data==True and flag_reject_target==True:
            fpr_r, tpr_r, threshold_r = roc_curve(preds_r,labels_r)  ###计算真正率和假正率
            roc_auc_r = auc(fpr_r, tpr_r)  ###计算auc的值  
        if flag_oot_data==True and flag_oot_target==True: 
            fpr_o, tpr_o, threshold_o = roc_curve(preds_o,labels_o)  ###计算真正率和假正率
            roc_auc_o = auc(fpr_o, tpr_o)  ###计算auc的值  
        gg=plt.figure()
        lw = 2
        plt.plot(fpr_t, tpr_t, color='darkorange',lw=lw , label='训练ROC curve (area = %0.3f)' % roc_auc_t)  ###假正率为横坐标，真正率为纵坐标做曲线
        if flag_v==True:
            plt.plot(fpr_v, tpr_v, color='grey',lw=lw ,linestyle='--', label='验证ROC curve (area = %0.3f)' % roc_auc_v)  ###假正率为横坐标，真正率为纵坐标做曲线
        if flag_reject_data==True and flag_reject_target==True:
            plt.plot(fpr_r, tpr_r, lw=lw ,linestyle='--', label='拒绝ROC curve (area = %0.3f)' % roc_auc_r)
        if flag_oot_data==True and flag_oot_target==True:
            plt.plot(fpr_o, tpr_o, lw=lw ,linestyle='--', label='OOT ROC curve (area = %0.3f)' % roc_auc_o)
        plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
        plt.xlim([0.0, 1.0])
        plt.ylim([0.0, 1.05])
        plt.xlabel('累计坏占比',fontsize=12)
        plt.ylabel('累计好占比',fontsize=12)
        plt.title('ROC')
        plt.legend(loc="lower right")
        plt.tight_layout()
        return gg
    def plotdis(self,df,target,score):
        try:
            plt.close()
        except:
            pass
        gg=plt.figure()
        n=plt.hist(df[df[target]==1][score], bins=100, density=1, alpha=0.5,
                  facecolor='blue',
                                     edgecolor='black',
                  histtype='bar',label='逾期客户'
                )
        m=plt.hist(df[df[target]==0][score], bins=100, density=1, alpha=0.5,
                  facecolor='green',
                                     edgecolor='black',
                  histtype='bar',label='正常客户'
                )
        m=plt.hist(df[score], bins=100, density=1, alpha=0.5,
                  facecolor='yellow',edgecolor='black',histtype='bar',label='所有客户')
        plt.legend()
        plt.title('训练集分数分布区间')
        plt.tight_layout()
        return gg
    def cali(self,df,df_v,score ,target,flag_v):
        try:
            plt.close()
        except:
            pass
        sp=self.num_finebin_group(df, score, 100,[])
        temp=self.binning(sp, df, score, 's_max', 's_min', 's_group', [])
        qq=temp.groupby([ 's_max', 's_min']).agg({score:'mean',target:'mean','s_group':'count'}).reset_index().rename({'s_group':'count'},axis=1)
        # QQ图和趋势线
        gg=plt.figure()
        z1 = np.polyfit(qq[score], qq[target], 1) # 用4次多项式拟合
        p1 = np.poly1d(z1)
        r=pow(qq[[score,target]].corr().iloc[1,0],2)
        note='训练集 y= %sx+%s'%(round(p1[1],3),round(p1[0],4)) +'\n R^2 = %s' %round(r,3)
         # 在屏幕上打印拟合多项式
        yvals=p1(qq[score]) # 也可以使用yvals=np.polyval(z1,x)
        
        if flag_v==True:
            
            temp_v=self.binning(sp, df_v, score, 's_max', 's_min', 's_group', [])
            qq_v=temp_v.groupby([ 's_max', 's_min']).agg({score:'mean',target:'mean','s_group':'count'}).reset_index().rename({'s_group':'count'},axis=1)
            # QQ图和趋势线

            z1_v = np.polyfit(qq_v[score], qq_v[target], 1) # 用4次多项式拟合
            p1_v = np.poly1d(z1_v)
            r_v=pow(qq_v[[score,target]].corr().iloc[1,0],2)
            note_v='\n验证集y= %sx+%s'%(round(p1_v[1],3),round(p1_v[0],4)) +'\n R^2 = %s' %round(r,3)
             # 在屏幕上打印拟合多项式
            note=note+note_v
            yvals_v=p1(qq_v[score]) # 也可以使用yvals=np.polyval(z1,x)            
            qq=pd.merge(qq,qq_v,how='left',on=['s_max','s_min'],suffixes=('','_v'))
        plot1=plt.plot(qq[score], qq[target], '*')
        plot2=plt.plot(qq[score], yvals, 'r',label='训练集趋势线')
        if flag_v==True:
            plot1=plt.plot(qq_v[score], qq_v[target], 'x')
            plot2=plt.plot(qq_v[score], yvals_v, 'b',label='验证集趋势线')
        plt.xlabel('分数',fontsize=12)
        plt.ylabel('坏账率',fontsize=12)
        plt.text(0, 0,note , fontsize=12, color = "r", style = "italic", weight = "light",alpha=0.5,
                 )
        plt.legend() # 指定legend的位置,读者可以自己help它的用法
        plt.title('Calibration')
        return qq, gg
    def binning(self,group_data, inputdata, col, inputmax, inputmin, inputgroup, specialcode_list):
        s_group_data = pd.DataFrame()
        group_data = group_data.reset_index()
        if specialcode_list != []:
            inputdata = inputdata[inputdata.isin({col: specialcode_list}) == False]
        inputdata = inputdata.loc[inputdata[col].isnull() == False]
        for lins in range(len(group_data)):
            temp = inputdata.copy()
            temp[inputgroup] = group_data.loc[lins, inputgroup]
            temp[inputmin] = group_data.loc[lins, inputmin]
            temp[inputmax] = group_data.loc[lins, inputmax]
            temp_data = temp[
                ((temp[col] <= temp[inputmax]) & (temp[col] > temp[inputmin]))
            ]
            s_group_data = pd.concat([s_group_data, temp_data])
        del temp
        return s_group_data
    def num_finebin_group(self,inputdata, col, s_bin_num, specialcode_list):
        if specialcode_list != []:
            inputdata = inputdata[inputdata.isin({col: specialcode_list}) == False]
        sort_df = inputdata[col][inputdata[col].isnull() == False]
        if len(sort_df.unique()) == 1:
            old_list = set([float('-inf'), float('inf')])
        elif len(sort_df.unique()) <= s_bin_num:
            old_list = set(list(sort_df.unique()))
            old_list.remove(max(old_list))
            old_list.remove(min(old_list))
            old_list.add(float('-inf'))
            old_list.add(float('inf'))
        else:
            old_list = set([float('-inf'), float('inf')])
            num = sort_df.size
            sort_df = sort_df.sort_values(ascending=True).reset_index().drop(columns='index')
            for i in range(1, s_bin_num):
                loca = int((i / s_bin_num) * num)
                value = sort_df.iloc[loca].values[0]
                old_list.add(value)
        new_list = list(old_list)
        new_list.sort()
        new_s_group = []
        for i in range(len(new_list) - 1):
            temp = {'s_group': i, 's_min': new_list[i], 's_max': new_list[i + 1]}
            new_s_group.append(temp)
        s_group_map = pd.DataFrame(new_s_group)
        return s_group_map
    def distimeall(self,df,df_v,df_t,df_r,
         oot_data,vaild_data,reject_data,score,train_time_id,oot_time_id,reject_time_id):
        try:
            plt.close()
        except:
            pass
        sp=self.num_finebin_group(df, score, 10,[])
        final_df=pd.DataFrame(columns=['time_id','s_max','s_min','s_group'])
        if train_time_id != None:
            temp=self.binning(sp, df, score, 's_max', 's_min', 's_group', [])
            qq=temp.groupby([train_time_id, 's_max', 's_min', 's_group']).agg({score:'count'}).reset_index().rename({train_time_id:'time_id',score:'count'},axis=1)
            qq_t = temp.groupby(train_time_id).agg({'s_group': 'count'}).reset_index().rename({train_time_id: 'time_id', 's_group': 'count_total'}, axis=1)
            qq=pd.merge(qq,qq_t,how='outer',on='time_id')
            qq['pct']=qq['count']/qq['count_total']
            final_df=pd.merge(final_df,qq,how='outer',on=['time_id','s_max','s_min', 's_group'])
        if vaild_data==True and train_time_id != None:
            temp=self.binning(sp, df_v, score, 's_max', 's_min', 's_group', [])
            qq=temp.groupby([train_time_id, 's_max', 's_min', 's_group']).agg({score:'count'}).reset_index().rename({train_time_id:'time_id',score:'count'},axis=1)
            qq_t = temp.groupby(train_time_id).agg({'s_group': 'count'}).reset_index().rename({train_time_id: 'time_id', 's_group': 'count_total'}, axis=1)
            qq=pd.merge(qq,qq_t,how='outer',on='time_id')
            qq['pct_v']=qq['count']/qq['count_total']
            final_df=pd.merge(final_df,qq,how='outer',on=['time_id','s_max','s_min', 's_group'])
        if reject_data == True and reject_time_id!=None:
            temp=self.binning(sp, df_r, score, 's_max', 's_min', 's_group', [])
            qq=temp.groupby([reject_time_id, 's_max', 's_min', 's_group']).agg({score:'count'}).reset_index().rename({reject_time_id:'time_id',score:'count'},axis=1)
            qq_t = temp.groupby(reject_time_id).agg({'s_group': 'count'}).reset_index().rename({reject_time_id: 'time_id', 's_group': 'count_total'}, axis=1)
            qq=pd.merge(qq,qq_t,how='outer',on='time_id')
            qq['pct_r']=qq['count']/qq['count_total']
            final_df=pd.merge(final_df,qq,how='outer',on=['time_id','s_max','s_min', 's_group'])
        if oot_data == True and oot_time_id != None:
            temp=self.binning(sp, df_t, score, 's_max', 's_min', 's_group', [])
            qq=temp.groupby([oot_time_id, 's_max', 's_min', 's_group']).agg({score:'count'}).reset_index().rename({oot_time_id:'time_id',score:'count'},axis=1)
            qq_t = temp.groupby(oot_time_id).agg({'s_group': 'count'}).reset_index().rename({oot_time_id: 'time_id', 's_group': 'count_total'}, axis=1)
            qq=pd.merge(qq,qq_t,how='outer',on='time_id')
            qq['pct_t']=qq['count']/qq['count_total']
            final_df=pd.merge(final_df,qq,how='outer',on=['time_id','s_max','s_min', 's_group'])
        final_df=final_df.fillna(0)
        #补上所有time_id 和 group 维度
        temp=final_df[['s_max', 's_min', 's_group']].drop_duplicates()
        plot_df=pd.DataFrame()
        for Time_id in set(final_df['time_id']):
            temp['time_id']=Time_id
            plot_df=plot_df.append(temp)
        plot_df=pd.merge(plot_df,final_df,how='left',on=['time_id','s_max','s_min', 's_group'])
        plot_df = plot_df.fillna(0)
        plot_df=plot_df.sort_values(by=['time_id','s_group'],ascending=[True,True])

        gg = plt.figure()
        ind = np.arange(len(set(plot_df['time_id'])))
        color_list=['gray','brown','orange','olive','green','cyan','blue','purple','pink','red','black']
        if train_time_id != None:
            BT = [0 for _ in range(len(set(plot_df['time_id'])))]
            c=0
            label_list = []
            pt = []
            for group in set(plot_df['s_group']):
                sort_df=plot_df[plot_df['s_group'] == group].sort_values(by='time_id',ascending=True)
                g=list(sort_df['pct'])
                x = list(sort_df['time_id'])
                s_min=round(list(plot_df[plot_df['s_group']==group]['s_min'])[0],3)
                s_max=round(list(plot_df[plot_df['s_group']==group]['s_max'])[0],3)
                p=plt.bar(ind, g, width=0.15, bottom=BT , color=color_list[c])
                label_list.append('%s-%s'%(s_min,s_max))
                pt.append(p)
                c=c+1
                d=[]
                for i in range(0, len(g)):
                    sum = BT[i] + g[i]
                    d.append(sum)
                BT=d
        if 'pct_v' in plot_df.columns:
            BT = [0 for _ in range(len(set(plot_df['time_id'])))]
            c = 0
            label_list = []
            pt = []
            for group in set(plot_df['s_group']):
                label_list = []
                sort_df = plot_df[plot_df['s_group'] == group].sort_values(by='time_id', ascending=True)
                g = list(sort_df['pct_v'])
                x = list(sort_df['time_id'])
                s_min=round(list(plot_df[plot_df['s_group']==group]['s_min'])[0],3)
                s_max=round(list(plot_df[plot_df['s_group']==group]['s_max'])[0],3)
                p=plt.bar(ind+(train_time_id != None)*0.155, g, width=0.15, bottom=BT,  color=color_list[c])
                label_list.append('%s-%s' % (s_min, s_max))
                pt.append(p)
                c = c + 1
                d=[]
                for i in range(0, len(g)):
                    sum = BT[i] + g[i]
                    d.append(sum)
                BT=d
        if 'pct_r' in plot_df.columns:
            BT = [0 for _ in range(len(set(plot_df['time_id'])))]
            c=0
            label_list = []
            pt = []
            for group in set(plot_df['s_group']):
                sort_df = plot_df[plot_df['s_group'] == group].sort_values(by='time_id', ascending=True)
                g = list(sort_df['pct_r'])
                x = list(sort_df['time_id'])
                s_min=round(list(plot_df[plot_df['s_group']==group]['s_min'])[0],3)
                s_max=round(list(plot_df[plot_df['s_group']==group]['s_max'])[0],3)
                p=plt.bar(ind+(train_time_id != None)*0.155+('pct_v' in plot_df.columns)*0.155, g, width=0.15, bottom=BT,color=color_list[c])
                label_list.append('%s-%s' % (s_min, s_max))
                pt.append(p)
                c = c + 1
                d=[]
                for i in range(0, len(g)):
                    sum = BT[i] + g[i]
                    d.append(sum)
                BT=d
        if 'pct_t' in plot_df.columns:
            BT = [0 for _ in range(len(set(plot_df['time_id'])))]
            c=0
            label_list = []
            pt = []
            for group in set(plot_df['s_group']):
                sort_df = plot_df[plot_df['s_group'] == group].sort_values(by='time_id', ascending=True)
                g = list(sort_df['pct_t'])
                x=list(sort_df['time_id'])
                s_min=round(list(plot_df[plot_df['s_group']==group]['s_min'])[0],3)
                s_max=round(list(plot_df[plot_df['s_group']==group]['s_max'])[0],3)
                p=plt.bar(ind+(train_time_id != None)*0.155+('pct_v' in plot_df.columns)*0.155+('pct_r' in plot_df.columns)*0.155, g, width=0.15, bottom=BT,color=color_list[c])
                label_list.append('%s-%s' % (s_min, s_max))
                pt.append(p)
                c = c + 1
                d=[]
                for i in range(0, len(g)):
                    sum = BT[i] + g[i]
                    d.append(sum)
                BT=d

        plt.ylabel('样本占比',fontsize=12)
        plt.title('时间分布图',fontsize=12)
        plt.xticks(ind, x ,rotation=30)
        # plt.legend()
        plt.legend([x[0] for x in pt], label_list,bbox_to_anchor=(-0.05, -0.15), loc="upper left", prop={'size': 9},
                   borderaxespad=0, ncol=5)
        plt.tight_layout()
        return gg
    def disallfb(self,df,df_v,df_t,df_r,
         oot_data,oot_target_flag,vaild_data,reject_data,reject_target_flag,
         train_target,reject_target,oot_target):
        try:
            plt.close()
        except:
            pass
        full=pd.DataFrame(columns=['plot_calibraition_flag'])
        full['plot_calibraition_flag']=('0.05','0.10','0.15','0.20','0.25','0.30','0.35','0.40','0.45','0.50','0.55','0.60','0.65','0.70','0.75','0.80', '0.85','0.90','0.95','1.00')

        def cal(x):
            for i in range(1,21):
                if x <= i*(1/20):
                    return '%4.2f' %(i*(1/20))
                    break
        df['plot_calibraition_flag'] = df['SCORECARD_LR_p_1'].apply(lambda x: cal(x)) 
        pre_plot=df.groupby('plot_calibraition_flag')[train_target].agg({'sum','count'}).reset_index()
        pre_plot['bad_pct']=pre_plot['sum']/pre_plot['count'].sum()
        pre_plot['total_pct']=pre_plot['count']/pre_plot['count'].sum()
        pre_plot['good_pct']=pre_plot['total_pct']-pre_plot['bad_pct']
        pre_plot=pd.merge(pre_plot,full,how='outer',on='plot_calibraition_flag').fillna(0)
        pre_plot=pre_plot.sort_values(by='plot_calibraition_flag')

        if vaild_data==True:
            df_v['plot_calibraition_flag'] = df_v['SCORECARD_LR_p_1'].apply(lambda x: cal(x)) 
            pre_plot_v=df_v.groupby('plot_calibraition_flag')[train_target].agg({'sum','count'}).reset_index()
            pre_plot_v['bad_pct']=pre_plot_v['sum']/pre_plot_v['count'].sum()
            pre_plot_v['total_pct']=pre_plot_v['count']/pre_plot_v['count'].sum()
            pre_plot_v['good_pct']=pre_plot_v['total_pct']-pre_plot_v['bad_pct']
            pre_plot_v=pd.merge(pre_plot_v,full,how='outer',on='plot_calibraition_flag').fillna(0)
            pre_plot_v = pre_plot_v.sort_values(by='plot_calibraition_flag')
        if oot_data==True:
            if oot_target_flag==True:
                df_t['plot_calibraition_flag'] = df_t['SCORECARD_LR_p_1'].apply(lambda x: cal(x)) 
                pre_plot_t=df_t.groupby('plot_calibraition_flag')[oot_target].agg({'sum','count'}).reset_index()
                pre_plot_t['bad_pct']=pre_plot_t['sum']/pre_plot_t['count'].sum()
                pre_plot_t['total_pct']=pre_plot_t['count']/pre_plot_t['count'].sum()
                pre_plot_t['good_pct']=pre_plot_t['total_pct']-pre_plot_t['bad_pct']
            else:
                df_t['plot_calibraition_flag'] = df_t['SCORECARD_LR_p_1'].apply(lambda x: cal(x)) 
                pre_plot_t=df_t.groupby('plot_calibraition_flag')['plot_calibraition_flag'].agg({'count'}).reset_index()
                pre_plot_t['total_pct']=pre_plot_t['count']/pre_plot_t['count'].sum()
            pre_plot_t=pd.merge(pre_plot_t,full,how='outer',on='plot_calibraition_flag').fillna(0)
            pre_plot_t = pre_plot_t.sort_values(by='plot_calibraition_flag')

        if reject_data==True:
            if reject_target_flag==True:
                df_r['plot_calibraition_flag'] = df_r['SCORECARD_LR_p_1'].apply(lambda x: cal(x)) 
                pre_plot_r=df_r.groupby('plot_calibraition_flag')[reject_target].agg({'sum','count'}).reset_index()
                pre_plot_r['bad_pct']=pre_plot_r['sum']/pre_plot_r['count'].sum()
                pre_plot_r['total_pct']=pre_plot_r['count']/pre_plot_r['count'].sum()
                pre_plot_r['good_pct']=pre_plot_r['total_pct']-pre_plot_r['bad_pct']
            else:
                df_r['plot_calibraition_flag'] = df_r['SCORECARD_LR_p_1'].apply(lambda x: cal(x)) 
                pre_plot_r=df_r.groupby('plot_calibraition_flag')['plot_calibraition_flag'].agg({'count'}).reset_index()
                pre_plot_r['total_pct']=pre_plot_r['count']/pre_plot_r['count'].sum()
            pre_plot_r=pd.merge(pre_plot_r,full,how='outer',on='plot_calibraition_flag').fillna(0)
            pre_plot_r = pre_plot_r.sort_values(by='plot_calibraition_flag')

        gg=plt.figure()
        N = len(pre_plot)
        GT = pre_plot['good_pct']
        BT = pre_plot['bad_pct']
        if vaild_data==True: 
            GV = pre_plot_v['good_pct']
            BV = pre_plot_v['bad_pct']

        if oot_data==True:
            if oot_target_flag==True:
                GO = pre_plot_t['good_pct']
                BO = pre_plot_t['bad_pct']
            else:
                oot=pre_plot_t['total_pct']
        if reject_data==True:
            if reject_target_flag==True:
                GR = pre_plot_r['good_pct']
                BR = pre_plot_r['bad_pct']
            else:
                reject=pre_plot_r['total_pct']

        ind = np.arange(N)    # the x locations for the groups
        width = 0.35       # the width of the bars: can also be len(x) sequence

        p1 = plt.bar(ind, GT, width=0.15,label='训练好客户' )#, yerr=menStd)
        p2 = plt.bar(ind, BT, width=0.15, bottom=GT,label='训练坏客户')#, yerr=womenStd)
        if vaild_data==True:
            p3 = plt.bar(ind+0.15, GV, width=0.15, label='验证好客户')#, yerr=menStd)
            p4 = plt.bar(ind+0.15, BV, width=0.15, bottom=GV,label='验证坏客户')#, yerr=womenStd)
        if oot_data==True:
            if oot_target_flag==True:
                p5 = plt.bar(ind+0.15* vaild_data+0.15, GO, width=0.15, label='OOT好客户')#, yerr=menStd)
                p6 = plt.bar(ind+0.15* vaild_data+0.15, BO, width=0.15, bottom=GO,label='OOT坏客户')#, yerr=womenStd)
            else:
                p7 = plt.bar(ind+0.15* vaild_data+0.15, oot, width=0.15,label='OOT客户')
        if reject_data==True:
            if reject_target_flag==True:
                p8 = plt.bar(ind+0.15* vaild_data+0.15* oot_data+0.15, GR, width=0.15, label='Reject好客户')#, yerr=menStd)
                p9 = plt.bar(ind+0.15* vaild_data+0.15* oot_data+0.15, BR, width=0.15, bottom=GR,label='Reject坏客户')#, yerr=womenStd)
            else:
                p10 = plt.bar(ind+0.15* vaild_data+0.15* oot_data+0.15, reject, width=0.15,label='Reject客户')
        plt.ylabel('样本占比',fontsize=12)
        plt.title('分数区间分布',fontsize=12)
        plt.xticks(ind, pre_plot['plot_calibraition_flag'])

        plt.legend()
        plt.tight_layout()
        return gg
    def groupplot_datainit(self,modify_var):

        grouped_traindata = self.predict_train_data
        grouped_validdata = self.predict_vaild_data
        grouped_rejectdata = self.predict_reject_data
        grouped_ootdata = self.predict_oot_data
        total_pct = pd.DataFrame(
            grouped_traindata.groupby(['f_group_%s' % modify_var])['f_group_%s' % modify_var].count()).rename(
            {'f_group_%s' % modify_var: 'num'}, axis=1).reset_index()
        total_pct['Group'] = total_pct['f_group_%s' % modify_var]
        total_pct['pct_train'] = total_pct['num'] / total_pct['num'].sum()
        total_pct = total_pct.drop(columns=['f_group_%s' % modify_var])

        bad_pct = grouped_traindata.groupby(['f_group_%s' % modify_var]).agg(
            {self.train_target: ['sum', 'count']}).reset_index().rename({'f_group_%s' % modify_var: 'Group'}, axis=1)
        bad_pct['badrate_train'] = bad_pct[self.train_target]['sum'] / bad_pct[self.train_target]['count']
        bad_pct = bad_pct[['Group', 'badrate_train']]

        if self.predict_vaild_data.empty==False:
            valid_pct = pd.DataFrame(grouped_validdata.groupby(['f_group_%s' % modify_var])[
                                         'f_group_%s' % modify_var].count()).rename(
                {'f_group_%s' % modify_var: 'num'}, axis=1).reset_index()
            valid_pct['Group'] = valid_pct['f_group_%s' % modify_var]
            valid_pct['pct_valid'] = valid_pct['num'] / valid_pct['num'].sum()
            valid_pct = valid_pct.drop(columns=['f_group_%s' % modify_var])
            total_pct = pd.merge(total_pct, valid_pct, how='outer', on='Group')

            validbad_pct = grouped_validdata.groupby(['f_group_%s' % modify_var]).agg(
                {self.train_target: ['sum', 'count']}).reset_index().rename({'f_group_%s' % modify_var: 'Group'}, axis=1)
            validbad_pct['badrate_valid'] = validbad_pct[self.train_target]['sum'] / validbad_pct[self.train_target]['count']
            validbad_pct = validbad_pct[['Group', 'badrate_valid']]
            bad_pct = pd.merge(bad_pct, validbad_pct, how='outer', on='Group')

        if self.predict_reject_data.empty==False and self.reject_target!=None and self.reject_time_id!=None :



            reject_pct = pd.DataFrame(grouped_rejectdata.groupby(['f_group_%s' % modify_var])[ 'f_group_%s' % modify_var].count()).rename(
                {'f_group_%s' % modify_var: 'num'}, axis=1).reset_index()
            reject_pct['Group'] = reject_pct['f_group_%s' % modify_var]
            reject_pct['pct_reject'] = reject_pct['num'] / reject_pct['num'].sum()
            reject_pct = reject_pct.drop(columns=['f_group_%s' % modify_var])
            total_pct = pd.merge(total_pct, reject_pct, how='outer', on='Group')
            if self.reject_target != None:
                rejectbad_pct = grouped_rejectdata.groupby(['f_group_%s' % modify_var]).agg(
                    {self.reject_target: ['sum', 'count']}).reset_index().rename({'f_group_%s' % modify_var: 'Group'}, axis=1)
                rejectbad_pct['badrate_reject'] = rejectbad_pct[self.reject_target]['sum'] / rejectbad_pct[self.reject_target]['count']
                rejectbad_pct = rejectbad_pct[['Group', 'badrate_reject']]
                bad_pct = pd.merge(bad_pct, rejectbad_pct, how='outer', on='Group')
        if self.predict_oot_data.empty==False and self.oot_target!=None and self.oot_time_id!=None :
            oot_pct = pd.DataFrame(grouped_ootdata.groupby(['f_group_%s' % modify_var])[
                                       'f_group_%s' % modify_var].count()).rename(
                {'f_group_%s' % modify_var: 'num'}, axis=1).reset_index()
            oot_pct['Group'] = oot_pct['f_group_%s' % modify_var]
            oot_pct['pct_oot'] = oot_pct['num'] / oot_pct['num'].sum()
            oot_pct = oot_pct.drop(columns=['f_group_%s' % modify_var])
            total_pct = pd.merge(total_pct, oot_pct, how='outer', on='Group')
            if self.oot_target != None:
                ootbad_pct = grouped_ootdata.groupby(['f_group_%s' % modify_var]).agg(
                    {self.oot_target: ['sum', 'count']}).reset_index().rename({'f_group_%s' % modify_var: 'Group'},axis=1)
                ootbad_pct['badrate_oot'] = ootbad_pct[self.oot_target]['sum'] / ootbad_pct[self.oot_target]['count']
                ootbad_pct = ootbad_pct[['Group', 'badrate_oot']]
                bad_pct = pd.merge(bad_pct, ootbad_pct, how='outer', on='Group')
#时间稳定性
        final=pd.DataFrame()
        if self.train_time_id != None:
            base_list = []
            for gr in grouped_traindata['f_group_%s' % modify_var].unique():
                for ti in grouped_traindata[self.train_time_id].unique():
                    base_list.append({'Group': gr, 'timeid': ti})
            base_df = pd.DataFrame(base_list)
            valid_pct = pd.DataFrame(
                grouped_traindata.groupby(['f_group_%s' % modify_var, self.train_time_id])[self.train_target].agg(
                    {'count', 'sum'}).reset_index())
            total = pd.DataFrame(
                grouped_traindata.groupby([self.train_time_id])[self.train_target].agg({'count'}).reset_index().rename(
                    {'count': 'total_num'}, axis=1))
            final = pd.merge(total, valid_pct, how='outer', on=self.train_time_id)
            final['popu_pct'] = final['count'] / final['total_num']
            final=final.rename({self.train_time_id:'timeid'},axis=1)
            final['bad_rate'] = final['sum'] / final['count']
            final['Group'] = final['f_group_%s' % modify_var]
            final = pd.merge(base_df, final, how='outer', on=['Group', 'timeid'])
        if (self.predict_vaild_data.empty==False)&(self.train_time_id != None):
            base_list=[]
            for gr in grouped_validdata['f_group_%s' % modify_var].unique():
                for ti in grouped_validdata[self.train_time_id].unique():
                    base_list.append({'Group':gr, 'timeid':ti})
            base_df=pd.DataFrame(base_list)
            valid_pct = pd.DataFrame(
                grouped_validdata.groupby(['f_group_%s' % modify_var, self.train_time_id])[self.train_target].agg(
                    {'count', 'sum'}).reset_index())
            total = pd.DataFrame(
                grouped_validdata.groupby([self.train_time_id])[self.train_target].agg({'count'}).reset_index().rename(
                    {'count': 'total_num'}, axis=1))
            final_v = pd.merge(total, valid_pct, how='outer', on=self.train_time_id)
            final_v['popu_pct'] = final_v['count'] / final_v['total_num']
            final_v['bad_rate'] = final_v['sum'] / final_v['count']
            final_v=final_v.rename({self.train_time_id:'timeid'},axis=1)
            final_v['Group'] = final_v['f_group_%s' % modify_var]
            final=pd.merge(final,final_v,how='outer',on=['Group','timeid'],suffixes=('','_valid'))
            final = pd.merge(base_df, final, how='outer', on=['Group', 'timeid'])
        if (self.predict_reject_data.empty==False) and (self.reject_time_id != None) :
            base_list=[]
            for gr in grouped_rejectdata['f_group_%s' % modify_var].unique():
                for ti in grouped_rejectdata[self.reject_time_id].unique():
                    base_list.append({'Group':gr, 'timeid':ti})
            base_df=pd.DataFrame(base_list)
            if self.reject_target != None:      
                valid_pct = pd.DataFrame(
                    grouped_rejectdata.groupby(['f_group_%s' % modify_var, self.reject_time_id])[
                        self.reject_target].agg({'count', 'sum'}).reset_index())
                total = pd.DataFrame(grouped_rejectdata.groupby([self.reject_time_id])[self.reject_target].agg(
                    {'count'}).reset_index().rename({'count': 'total_num'}, axis=1))
                final_r = pd.merge(total, valid_pct, how='outer', on=self.reject_time_id)
                final_r['popu_pct'] = final_r['count'] / final_r['total_num']
                final_r['bad_rate'] = final_r['sum'] / final_r['count']
                final_r=final_r.rename({self.reject_time_id:'timeid'},axis=1)
                final_r['Group'] = final_r['f_group_%s' % modify_var]
                final=pd.merge(final,final_r,how='outer',on=['Group','timeid'],suffixes=('','_reject'))
                final = pd.merge(base_df, final, how='outer', on=['Group', 'timeid'])
            else:
                valid_pct = pd.DataFrame(
                    grouped_rejectdata.groupby(['f_group_%s' % modify_var, self.reject_time_id])[
                        modify_var].agg({'count'}).reset_index())
                total = pd.DataFrame(
                    grouped_rejectdata.groupby([self.reject_time_id])['f_group_%s' % modify_var].agg(
                        {'count'}).reset_index().rename({'count': 'total_num'}, axis=1))
                final_r = pd.merge(total, valid_pct, how='outer', on=self.reject_time_id)
                final_r['popu_pct'] = final_r['count'] / final_r['total_num']
                final_r=final_r.rename({self.reject_time_id:'timeid'},axis=1)
                final_r['Group'] = final_r['f_group_%s' % modify_var]
                final=pd.merge(final,final_r,how='outer',on=['Group','timeid'],suffixes=('','_reject'))
                final = pd.merge(base_df, final, how='outer', on=['Group', 'timeid'])
        if self.predict_oot_data.empty==False  and self.oot_time_id!=None :
            base_list=[]
            for gr in grouped_ootdata['f_group_%s' % modify_var].unique():
                for ti in grouped_ootdata[self.oot_time_id].unique():
                    base_list.append({'Group':gr,'timeid':ti})
            base_df=pd.DataFrame(base_list)
            if self.oot_target != None:
                valid_pct = pd.DataFrame(
                    grouped_ootdata.groupby(['f_group_%s' % modify_var, self.oot_time_id])[self.oot_target].agg(
                        {'count', 'sum'}).reset_index())
                total = pd.DataFrame(
                    grouped_ootdata.groupby([self.oot_time_id])[self.oot_target].agg({'count'}).reset_index().rename(
                        {'count': 'total_num'}, axis=1))
                final_o = pd.merge(total, valid_pct, how='outer', on=self.oot_time_id)
                final_o['popu_pct'] = final_o['count'] / final_o['total_num']
                final_o['bad_rate'] = final_o['sum'] / final_o['count']
                final_o=final_o.rename({self.oot_time_id:'timeid'},axis=1)
                final_o['Group'] = final_o['f_group_%s' % modify_var]
                final=pd.merge(final,final_o,how='outer',on=['Group','timeid'],suffixes=('','_oot'))
                final = pd.merge(base_df, final, how='outer', on=['Group', 'timeid'])
            else:
                valid_pct = pd.DataFrame(
                    grouped_ootdata.groupby(['f_group_%s' % modify_var, self.oot_time_id])[modify_var].agg(
                        {'count'}).reset_index())
                total = pd.DataFrame(grouped_ootdata.groupby([self.oot_time_id])['f_group_%s' % modify_var].agg(
                    {'count'}).reset_index().rename({'count': 'total_num'}, axis=1))
                final_o = pd.merge(total, valid_pct, how='outer', on=self.oot_time_id)
                final_o['popu_pct'] = final_o['count'] / final['total_num']
                final_o=final_o.rename({self.oot_time_id:'timeid'},axis=1)
                final_o['Group'] = final_o['f_group_%s' % modify_var]
                final=pd.merge(final,final_o,how='outer',on=['Group','timeid'],suffixes=('','_oot'))
                final = pd.merge(base_df, final, how='outer', on=['Group', 'timeid'])
        return bad_pct, total_pct, final
    def plot_group(self, bad_pct, total_pct, final,directtion='V'):
        try:
            plt.close()
        except:
            pass
        total_pic=1
        if final.empty==False:
            final=final.sort_values(by='timeid',ascending=True)
            if  'popu_pct'  in list(final.columns) :
                total_pic=total_pic+1
            if  'popu_pct_valid'  in list(final.columns) :
                total_pic=total_pic+1
            if 'popu_pct_reject'  in list(final.columns) :
                total_pic=total_pic+1
            if 'popu_pct_oot'  in list(final.columns) :
                total_pic=total_pic+1
        if directtion == 'V':
            fig = plt.figure(figsize=(10,6.3*total_pic),dpi=int(((self.screenwidth*0.4)/10)))
        else:
            fig = plt.figure(figsize=( 10 * total_pic,5.5), dpi=int(((self.screenwidth*0.4)/ 10)))
        
#      
        plt.subplots_adjust(left=None, bottom=None, right=None, top=None,
                            wspace=0.8, hspace=0.3)
        num_pic=1

    #-----------------------------------------------------
        # 组分布
        if directtion == 'V':
            ax1 = fig.add_subplot(total_pic, 1, num_pic)
        else:
            ax1 = fig.add_subplot(1,total_pic, num_pic)
        x = total_pct['Group']
        y1 = total_pct['pct_train']
        bar_width = 0.2
        plt.bar(x=range(len(x)), height=y1, label='Train', color='steelblue',  width=bar_width)
        if 'pct_valid' in list(total_pct.columns):
            y2 = total_pct['pct_valid']
            plt.bar(x=np.arange(len(x)) + bar_width, height=y2, label='valid', color='indianred', 
                    width=bar_width)
        if 'pct_reject' in list(total_pct.columns):
            y3 = total_pct['pct_reject']
            plt.bar(x=np.arange(len(x)) + 2 * bar_width, height=y3, label='reject', color='powderblue', 
                    width=bar_width)
        if 'pct_oot' in list(total_pct.columns):
            y4 = total_pct['pct_oot']
            plt.bar(x=np.arange(len(x)) + 3 * bar_width, height=y4, label='oot', color='pink', 
                    width=bar_width)

        plt.legend(loc='upper center',ncol=5)
        plt.xlabel(u"组",fontsize=12)
        plt.ylabel(u"样本占比",fontsize=12)
        plt.title('分组分布',fontsize=12)
        badrate1 = bad_pct['badrate_train']
        ax2 = ax1.twinx()
        ax2.plot(np.arange(len(x)), badrate1, 'steelblue')
        if ('badrate_valid', '') in list(bad_pct.columns):
            badrate2 = bad_pct['badrate_valid']
            ax2.plot(np.arange(len(x)) + bar_width, badrate2, 'indianred')
        if ('badrate_reject', '') in list(bad_pct.columns):
            badrate3 = bad_pct['badrate_reject']
            ax2.plot(np.arange(len(x)) + bar_width * 2, badrate3, 'powderblue')
        if ('badrate_oot', '') in list(bad_pct.columns):
            badrate4 = bad_pct['badrate_oot']
            ax2.plot(np.arange(len(x)) + bar_width * 3, badrate4, 'pink')
        ax2.set_ylabel('坏账率',fontsize=12)
        plt.xticks(np.arange(len(x)) + 0.15, x, fontsize=12, rotation=20)
        num_pic=num_pic+1
    #-----------------------------------------------------
        # Train

        if 'popu_pct'  in list(final.columns) :
            if directtion == 'V':
                ax3 = fig.add_subplot(total_pic, 1, num_pic)
            else:
                ax3 = fig.add_subplot(1,total_pic, num_pic)
            # ax3 = fig.add_subplot(total_pic, 1, num_pic)
            bar_width = 0.1
            i = 0
            group_list = list(final['Group'].unique())
            group_list.sort(reverse=False)
            for group in group_list:
                tt = final[final['Group'] == group]
                y2 = tt['popu_pct']
                x = tt['timeid']
                plt.bar(x=np.arange(len(x)) + 0.08 * i, height=y2, label="组_%s" % group,
                        width=bar_width)
                i = i + 1
            plt.legend(loc='upper center',ncol=5)

            plt.xlabel("Time",fontsize=12)
            plt.ylabel("样本占比",fontsize=12)
            plt.title('Train',fontsize=12)
            plt.xticks(np.arange(len(x)) + 0.15, x, fontsize=12, rotation=45)
            if 'bad_rate' in list(tt.columns):
                ax4 = ax3.twinx()  # this is the important function
                i = 0
                group_list = list(final['Group'].unique())
                group_list.sort(reverse=False)
                for group in group_list:
                    tt = final[final['Group'] == group]
                    y = tt['bad_rate']
                    x = tt['timeid']
                    ax4.plot(np.arange(len(x)) + 0.08 * i, y)
                    i = i + 1
                ax4.set_ylabel('坏账率',fontsize=12)
            num_pic=num_pic+1
    #-----------------------------------------------------
        # vaild        
        if 'popu_pct_valid'  in list(final.columns) :
            # Train
            if directtion == 'V':
                ax4 = fig.add_subplot(total_pic, 1, num_pic)
            else:
                ax4 = fig.add_subplot(1,total_pic, num_pic)
            # ax4 = fig.add_subplot(total_pic, 1, num_pic)
            bar_width = 0.1
            i = 0
            group_list = list(final['Group'].unique())
            group_list.sort(reverse=False)
            for group in group_list:
                tt = final[final['Group'] == group]
                y2 = tt['popu_pct_valid']
                x = tt['timeid']
                plt.bar(x=np.arange(len(x)) + 0.08 * i, height=y2, label="组_%s" % group,  
                        width=bar_width)
                i = i + 1
            plt.legend(loc='upper center',ncol=5)

            plt.xlabel("Time",fontsize=12)
            plt.ylabel("样本占比",fontsize=12)
            plt.title('Valid',fontsize=12)
            plt.xticks(np.arange(len(x)) + 0.15, x, fontsize=12, rotation=45)
            if 'bad_rate_valid' in list(tt.columns):
                ax5 = ax4.twinx()  # this is the important function
                i = 0
                for group in list(final['Group'].unique()):
                    tt = final[final['Group'] == group]
                    y = tt['bad_rate_valid']
                    x = tt['timeid']
                    ax5.plot(np.arange(len(x)) + 0.08 * i, y)
                    i = i + 1
                ax5.set_ylabel('坏账率',fontsize=12)
            num_pic=num_pic+1
    #-----------------------------------------------------
        # reject 
        if 'popu_pct_reject'  in list(final.columns) :
            if directtion == 'V':
                ax6 = fig.add_subplot(total_pic, 1, num_pic)
            else:
                ax6 = fig.add_subplot(1,total_pic, num_pic)
            # ax6 = fig.add_subplot(total_pic, 1, num_pic)
            bar_width = 0.1
            i = 0
            group_list=list(final['Group'].unique())
            group_list.sort(reverse=False)
            for group in group_list:
                tt = final[final['Group'] == group]
                y2 = tt['popu_pct_reject']
                x = tt['timeid']
                plt.bar(x=np.arange(len(x)) + 0.08 * i, height=y2, label="组_%s" % group,  
                        width=bar_width)
                i = i + 1
            plt.legend(loc='upper center',ncol=5)

            plt.xlabel("Time",fontsize=12)
            plt.ylabel("样本占比",fontsize=12)
            plt.title('Reject',fontsize=12)
            plt.xticks(np.arange(len(x)) + 0.15, x, fontsize=12, rotation=45)
            if 'bad_rate_reject' in list(tt.columns):
                ax7 = ax6.twinx()  # this is the important function
                i = 0
                for group in list(final['Group'].unique()):
                    tt = final[final['Group'] == group]
                    y = tt['bad_rate_reject']
                    x = tt['timeid']
                    ax7.plot(np.arange(len(x)) + 0.08 * i, y)
                    i = i + 1
                ax7.set_ylabel('坏账率',fontsize=12)
            num_pic=num_pic+1

    #-----------------------------------------------------
        # oot 
        if 'popu_pct_oot'  in list(final.columns) :
            if directtion == 'V':
                ax8 = fig.add_subplot(total_pic, 1, num_pic)
            else:
                ax8 = fig.add_subplot(1,total_pic, num_pic)
            # ax8 = fig.add_subplot(total_pic, 1, num_pic)
            bar_width = 0.1
            i = 0
            group_list = list(final['Group'].unique())
            group_list.sort(reverse=False)
            for group in group_list:
                tt = final[final['Group'] == group]
                y2 = tt['popu_pct_oot']
                x = tt['timeid']
                plt.bar(x=np.arange(len(x)) + 0.08 * i, height=y2, label="组_%s" % group,  
                        width=bar_width)
                i = i + 1
            plt.legend(loc='upper center',ncol=5)

            plt.xlabel("Time",fontsize=12)
            plt.ylabel("样本占比",fontsize=12)
            plt.title('OOT',fontsize=12)
            plt.xticks(np.arange(len(x)) + 0.15, x, fontsize=12, rotation=45)
            if 'bad_rate_oot' in list(tt.columns):
                ax9= ax8.twinx()  # this is the important function
                i = 0
                for group in list(final['Group'].unique()):
                    tt = final[final['Group'] == group]
                    y = tt['bad_rate_oot']
                    x = tt['timeid']
                    ax9.plot(np.arange(len(x)) + 0.08 * i, y)
                    i = i + 1
                ax9.set_ylabel('坏账率',fontsize=12)
        plt.tight_layout()
        return fig


    def calculateginitime(self,gini_score, predict_train_data, train_time_id,train_target,flag_v,predict_vaild_data,
                          predict_oot_data,oot_target,oot_time_id,predict_reject_data,reject_time_id,reject_target):
        ste_da = pd.DataFrame()
        if train_time_id != None:
            for period in list(predict_train_data[train_time_id].unique()):
                fpr, tpr, threshold = roc_curve(
                    predict_train_data[predict_train_data[train_time_id] == period][train_target]
                    , predict_train_data[predict_train_data[train_time_id] == period][gini_score])
                roc_auc = auc(fpr, tpr)
                num = predict_train_data[predict_train_data[train_time_id] == period][train_time_id].count()
                bad_rate = predict_train_data[predict_train_data[train_time_id] == period][train_target].mean()
                dd = {'timeid': period, 'auc': roc_auc, 'count': num, 'bad_rate': bad_rate}
                ste_da = ste_da.append(pd.DataFrame(dd, index=[1]))
                ste_da['auc'] = round(ste_da['auc'], 3)
                ste_da['bad_rate'] = round(ste_da['bad_rate'], 3)
        if (flag_v == True) and (train_time_id != None):
            temp_v = pd.DataFrame()
            for period in list(predict_vaild_data[train_time_id].unique()):
                fpr, tpr, threshold = roc_curve(
                    predict_vaild_data[predict_vaild_data[train_time_id] == period][train_target]
                    , predict_vaild_data[predict_vaild_data[train_time_id] == period][gini_score])  ###计算真正率和假正率
                roc_auc = auc(fpr, tpr)
                num = predict_vaild_data[predict_vaild_data[train_time_id] == period][train_target].count()
                bad_rate = predict_vaild_data[predict_vaild_data[train_time_id] == period][train_target].mean()
                dd = {'timeid': period, 'auc': roc_auc, 'count': num, 'bad_rate': bad_rate}
                temp_v = temp_v.append(pd.DataFrame(dd, index=[1]))
                temp_v['auc'] = round(temp_v['auc'], 3)
                temp_v['bad_rate'] = round(temp_v['bad_rate'], 3)
            ste_da = pd.merge(ste_da, temp_v, how='outer', on=['timeid'], suffixes=('', '_v'))

        if (predict_oot_data.empty == False) and (oot_target != None) and (oot_time_id != None):
            temp_o = pd.DataFrame()
            for period in list(predict_oot_data[oot_time_id].unique()):
                fpr, tpr, threshold = roc_curve(predict_oot_data[predict_oot_data[oot_time_id] == period][oot_target]
                                                , predict_oot_data[predict_oot_data[oot_time_id] == period][
                                                    gini_score])  ###计算真正率和假正率
                roc_auc = auc(fpr, tpr)
                num = predict_oot_data[predict_oot_data[oot_time_id] == period][oot_target].count()
                bad_rate = predict_oot_data[predict_oot_data[oot_time_id] == period][oot_target].mean()
                dd = {'timeid': period, 'auc': roc_auc, 'count': num, 'bad_rate': bad_rate}
                temp_o = temp_o.append(pd.DataFrame(dd, index=[1]))
                temp_o['auc'] = round(temp_o['auc'], 3)
                temp_o['bad_rate'] = round(temp_o['bad_rate'], 3)
            ste_da = pd.merge(ste_da, temp_o, how='outer', on=['timeid'], suffixes=('', '_oot'))

        if predict_reject_data.empty == False and reject_target != None and reject_time_id != None:
            temp_r = pd.DataFrame()
            for period in list(predict_reject_data[reject_time_id].unique()):
                fpr, tpr, threshold = roc_curve(
                    predict_reject_data[predict_reject_data[reject_time_id] == period][reject_target]
                    , predict_reject_data[predict_reject_data[reject_time_id] == period][gini_score])  ###计算真正率和假正率
                roc_auc = auc(fpr, tpr)
                num = predict_reject_data[predict_reject_data[reject_time_id] == period][reject_target].count()
                bad_rate = predict_reject_data[predict_reject_data[reject_time_id] == period][reject_target].mean()
                dd = {'timeid': period, 'auc': roc_auc, 'count': num, 'bad_rate': bad_rate}
                temp_r = temp_r.append(pd.DataFrame(dd, index=[1]))
                temp_r['auc'] = round(temp_r['auc'], 3)
                temp_r['bad_rate'] = round(temp_r['bad_rate'], 3)
            ste_da = pd.merge(ste_da, temp_r, how='outer', on=['timeid'], suffixes=('', '_reject'))
        if ste_da.empty == False:
            ste_da = ste_da.sort_values(by='timeid', ascending=False)
        return ste_da

    def output_report(self):
        alphabet = '1ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        def col_num_alp(col_num):
            sec_num = col_num // 26
            first_num = col_num % 26
            if first_num == 0:
                sec_num = sec_num - 1
                first_num = 26
            al_col = (sec_num != 0) * alphabet[sec_num] + alphabet[first_num]
            return al_col
        def boder_set(ws,start_col,end_col,start_row,end_row,tile=False):
            alphabet = '1ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            for col_num in range(start_col,end_col+1):
                for row_num in range(start_row,end_row+1):
                    if col_num==start_col:
                        left_boder=BORDER_THIN
                    else:
                        left_boder = None
                    if col_num==end_col:
                        right_boder=BORDER_THIN
                    else:
                        right_boder = None
                    if row_num==start_row:
                        top_boder=BORDER_THIN
                    else:
                        top_boder = None
                    if row_num==end_row:
                        bottom_boder=BORDER_THIN
                    else:
                        bottom_boder=None
                    border_setting = Border(
                        left=Side(border_style=left_boder, color=colors.BLACK),
                        right=Side(border_style=right_boder, color=colors.BLACK),
                        top=Side(border_style=top_boder, color=colors.BLACK),
                        bottom=Side(border_style=bottom_boder, color=colors.BLACK)
                    )

                    sec_num = col_num // 26
                    first_num = col_num % 26
                    if first_num==0:
                        sec_num=sec_num-1
                        first_num=26
                    al_col = (sec_num != 0) * alphabet[sec_num] + alphabet[first_num]
                    ws['%s%d' % (al_col,row_num)].border = border_setting
                    if tile==True:
                        ws['%s%d' % (al_col,row_num)].fill = PatternFill("solid", fgColor='6E8B3D')
        wb= Workbook()
        #第一页
        workSheet = wb.create_sheet('模型信息')
        sc = self.scorecarddf[['variable_name', 'f_group', 'label', 'f_Bad_rate', 'pct_f_N_obs', 'woe', 'coff', 'scorecard']].rename(
                                {"variable_name": "变量名称",
                                 'f_group': "分组",
                                 'label': "注释",
                                 'f_Bad_rate': "坏账率",
                                 'pct_f_N_obs': "样本占比",
                                 'woe': "WOE",
                                 'coff': "系数",
                                 'scorecard': "分数"}, axis=1)
        z =self.model.summary()
        m = z.as_text()

        # 模型信息
        # workSheet.title = '模型信息'
        model_info_col_num = math.ceil(max([len(str(x)) for x in list(self.scorecarddf['variable_name'])]) / 8)+9
        summary_num=col_num_alp(model_info_col_num)+str(m.count('\n') + 2)
        workSheet.merge_cells('A2:%s' % summary_num)
        workSheet.cell(1, 1).value = '模型信息：'
        boder_set(workSheet, start_col=1, end_col=1, start_row=1, end_row=1,tile=1)
        workSheet.cell(2, 1).value = m
        font = Font(name=u'宋体', bold=True)
        align = Alignment(horizontal='left', vertical='top', wrapText=True)
        workSheet['A2'].alignment = align
        boder_set(workSheet, start_col=2, end_col=model_info_col_num, start_row=2,end_row=m.count('\n') + 2, tile=False)

        #变量重要性
        workSheet.cell(1, model_info_col_num+3).value = '变量重要性：'
        boder_set(workSheet, start_col=model_info_col_num+3, end_col=model_info_col_num+3, start_row=1,end_row=1, tile=1)
        imgdata = io.BytesIO()
        self.imp_gg.savefig(imgdata, format='png')
        img = openpyxl.drawing.image.Image(imgdata)
        img.anchor='%s2'%col_num_alp(model_info_col_num+3)
        workSheet.add_image(img)
        # 评分卡信息
        workSheet.cell(m.count('\n') + 3, 1).value = '评分卡：'
        boder_set(workSheet, start_col=1, end_col=1, start_row=m.count('\n') + 3, end_row=m.count('\n') + 3,tile=1)
        name_num = math.ceil(max([len(str(x)) for x in list(self.scorecarddf['variable_name'])]) / 8)
        name_col = col_num_alp(name_num)
        label_col_start= col_num_alp(name_num+2)
        label_num = math.ceil(max([len(str(x)) for x in list(self.scorecarddf['label'])]) / 8)
        label_num = name_num + 1 + label_num
        label_col = col_num_alp(label_num)
        rows = dataframe_to_rows(sc, index=False)
        for r_idx, row in enumerate(rows, 1):
            workSheet.merge_cells('A%d:%s%d' % (r_idx + m.count('\n') + 4, name_col, r_idx + m.count('\n') + 4))
            workSheet.merge_cells('%s%d:%s%d' % (label_col_start, r_idx + m.count('\n') + 4, label_col, r_idx + m.count('\n') + 4))
            if r_idx == 1:
                title_flag=True
            else:
                title_flag = False
            for c_idx, value in enumerate(row, 1):
                if c_idx == 1:
                    workSheet.cell(row=r_idx + m.count('\n') + 4, column=c_idx, value=value)
                    boder_set(workSheet,start_col=1,end_col=name_num,start_row=r_idx + m.count('\n') + 4,end_row=r_idx + m.count('\n') + 4,tile=title_flag)
                elif c_idx == 2:
                    workSheet.cell(row=r_idx + m.count('\n') + 4, column=name_num + 1, value=value)
                    boder_set(workSheet,start_col=name_num + 1,end_col=name_num + 1,start_row=r_idx + m.count('\n') + 4,end_row=r_idx + m.count('\n') + 4,tile=title_flag)
                elif c_idx == 3:
                    workSheet.cell(row=r_idx + m.count('\n') + 4, column=name_num + 2, value=value)
                    boder_set(workSheet,start_col=name_num + 2,end_col=label_num,start_row=r_idx + m.count('\n') + 4,end_row=r_idx + m.count('\n') + 4,tile=title_flag)
                else:
                    workSheet.cell(row=r_idx + m.count('\n') + 4, column=label_num + c_idx-3, value=value)
                    boder_set(workSheet,start_col=(label_num +c_idx -3),end_col=(label_num +c_idx-3),start_row=(r_idx + m.count('\n') + 4),end_row=(r_idx + m.count('\n') + 4),tile=title_flag)

        # 模型筛选变量信息

        # recordlist = ""
        # for re in self.record_list:
        #     if isinstance(re, list):
        #         for ree in re:
        #             recordlist = recordlist + str(ree) + '\n'
        #     elif isinstance(re, str):
        #         recordlist = recordlist + re + '\n'
        #     else:
        #         recordlist = recordlist + re.as_text()
        # start = 0
        # str_list = []
        # for i in range(0, len(recordlist)):
        #     if recordlist[i:i + 1] == '\n':
        #         end = i + 1
        #         str_list.append(recordlist[start:end - 1])
        #         start = end
        #
        # relist_col_satr = col_num_alp(label_num+8)
        # relist_col_end= col_num_alp(label_num + 16)
        # workSheet.cell(m.count('\n') + 3, label_num + 8).value = '模型记录：'
        # # workSheet.merge_cells('%s%d:%s%d' % (relist_col_satr, m.count('\n') + 4, relist_col_end, m.count('\n') + 3+recordlist.count('\n')+8))
        # # workSheet.cell(m.count('\n') + 4,label_num+8).value = recordlist
        # # align = Alignment(horizontal='left', vertical='top', wrapText=True)
        # # workSheet['%s%d' % (col_num_alp(label_num+8), m.count('\n') + 4)].alignment = align
        # # boder_set(workSheet, start_col=label_num+8, end_col=label_num*2 + 16, start_row=m.count('\n') + 4,end_row=m.count('\n') + 3+recordlist.count('\n')+8, tile=False)
        # for row_num in range(0, len(str_list)):
        #     print(relist_col_satr, row_num, relist_col_end, row_num)
        #     workSheet.merge_cells('%s%d:%s%d' % (relist_col_satr, m.count('\n') + 4+row_num, relist_col_end,  m.count('\n') + 4+row_num))
        #     workSheet.cell(m.count('\n') + 4+row_num, label_num + 8).value = str_list[row_num ]
        #     font = Font(name=u'宋体', bold=True)
        #     align = Alignment(horizontal='left', vertical='top', wrapText=True)
        #     workSheet['%s%d' %(relist_col_satr, m.count('\n') + 4+row_num)].alignment = align
        # boder_set(workSheet, start_col=label_num+8, end_col=label_num + 16, start_row=m.count('\n') + 4,
        #           end_row=m.count('\n') + 4+len(str_list), tile=title_flag)

        #第二页

        workSheet2 = wb.create_sheet('模型表现')
        # workSheet2.title = '模型表现'
        #ks
        workSheet2.cell(1, 1).value = '模型KS表现：'
        boder_set(workSheet2, start_col=1, end_col=1, start_row=1,end_row=1, tile=1)
        imgdata = io.BytesIO()
        self.ks_mp.savefig(imgdata, format='png')
        img = openpyxl.drawing.image.Image(imgdata)
        img.anchor='A2'
        workSheet2.add_image(img)
        # AUC
        workSheet2.cell(31, 1).value = '模型AUC表现：'
        boder_set(workSheet2, start_col=1, end_col=1, start_row=31,end_row=31, tile=1)
        imgdata = io.BytesIO()
        self.auc_m.savefig(imgdata, format='png')
        img = openpyxl.drawing.image.Image(imgdata)
        img.anchor='A32'
        workSheet2.add_image(img)
        #lift
        workSheet2.cell(61, 1).value = '模型Lift表现：'
        boder_set(workSheet2, start_col=1, end_col=1, start_row=61,end_row=61, tile=1)
        imgdata = io.BytesIO()
        self.lift_m.savefig(imgdata, format='png')
        img = openpyxl.drawing.image.Image(imgdata)
        img.anchor='A62'
        workSheet2.add_image(img)
        #分数分布
        workSheet2.cell(91, 1).value = '模型分数分布：'
        boder_set(workSheet2, start_col=1, end_col=1, start_row=91,end_row=91, tile=1)
        imgdata = io.BytesIO()
        self.dis_train_m.savefig(imgdata, format='png')
        img = openpyxl.drawing.image.Image(imgdata)
        img.anchor='A92'
        workSheet2.add_image(img)
        #多样本分数分布
        workSheet2.cell(121, 1).value = '多样本分数分布：'
        boder_set(workSheet2, start_col=1, end_col=1, start_row=121,end_row=121, tile=1)
        imgdata = io.BytesIO()
        self.all_sample_dis_m.savefig(imgdata, format='png')
        img = openpyxl.drawing.image.Image(imgdata)
        img.anchor='A122'
        workSheet2.add_image(img)
        #分数校准
        workSheet2.cell(151, 1).value = '分数校准：'
        boder_set(workSheet2, start_col=1, end_col=1, start_row=151,end_row=151, tile=1)
        imgdata = io.BytesIO()
        self.cail_m.savefig(imgdata, format='png')
        img = openpyxl.drawing.image.Image(imgdata)
        img.anchor='A152'
        workSheet2.add_image(img)
        # 分数校准数据
        workSheet2.cell(181, 1).value = '分数校准数据：'
        boder_set(workSheet2, start_col=1, end_col=1, start_row=181, end_row=181, tile=1)
        if self.predict_vaild_data.empty==False:
            out_cail_data=self.cailbra_data[['s_group','SCORECARD_LR_p_1',self.train_target, 'count','SCORECARD_LR_p_1_v', self.train_target+'_v', 'count_v']].rename({
                's_group':"分数段",'SCORECARD_LR_p_1': "Trian预测值",self.train_target: "Trian坏账率",'count': "Trian样本数",'SCORECARD_LR_p_1_v': "vaild预测值",
                self.train_target + '_v':"vaild预测值坏账率",'count_v': "vaild预测值样本数"},axis=1)
        else:
            out_cail_data=self.cailbra_data[['s_group', 'SCORECARD_LR_p_1', self.train_target, 'count']].rename({
                's_group': "分数段", 'SCORECARD_LR_p_1': "Trian预测值", self.train_target: "Trian坏账率", 'count': "Trian样本数"},axis=1)
        rows = dataframe_to_rows(out_cail_data, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                workSheet2.cell(row=r_idx + 183, column=c_idx, value=value)
                if r_idx==1:
                    boder_set(workSheet2, start_col=c_idx, end_col=c_idx, start_row=r_idx + 183,end_row=r_idx + 183, tile=True)
                # lif数据
        workSheet2.cell(181, len(out_cail_data.columns)+3).value = 'lift数据：'
        boder_set(workSheet2, start_col=len(out_cail_data.columns)+3, end_col=len(out_cail_data.columns)+3, start_row=181, end_row=181, tile=1)
        if self.predict_vaild_data.empty==False:
            out_lift_data=self.lift_data[['total_pct', 'lift','total_pct_v', 'lift_v']].rename({
                                                                                'total_pct':"Train样本累计占比",
                                                                                'lift':"Train提升率",
                                                                                'total_pct_v':"Valid样本累计占比",
                                                                                'lift_v':"Vliad提升率"},axis=1)
        else:
            out_lift_data=self.lift_data[['total_pct', 'lift']].rename({
                                                        'total_pct': "Train样本累计占比",
                                                        'lift': "Train提升率"}, axis=1)
        rows = dataframe_to_rows(out_lift_data, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                workSheet2.cell(row=r_idx + 183, column=c_idx+len(out_cail_data.columns)+2, value=value)
                if r_idx==1:
                    boder_set(workSheet2, start_col=c_idx+len(out_cail_data.columns)+2, end_col=c_idx+len(out_cail_data.columns)+2, start_row=r_idx + 183,end_row=r_idx + 183, tile=True)
    #第三页
        if (self.train_time_id != None) or (self.reject_time_id != None and self.predict_reject_data.empty == False) or (
                self.oot_time_id != None and self.predict_oot_data.empty == False):
            workSheet3 = wb.create_sheet('模型时间稳定性')
            workSheet3.cell(1, 1).value = '模型稳定性表现：'
            boder_set(workSheet3, start_col=1, end_col=1, start_row=1, end_row=1, tile=1)

            if self.model_ginitime.empty==False:
                col=["时间ID"]
                data_df=self.model_ginitime.copy()
                data_df=data_df.rename({'timeid': "时间ID"},axis=1)
                if 'count' in list(self.model_ginitime.columns):
                    data_df=data_df.rename({
                                            'auc':"Trian AUC",
                                            'count': "Trian样本数",
                                            'bad_rate': "Trian坏账率"},axis=1)
                    col=col+["Trian AUC","Trian样本数","Trian坏账率"]
                if 'count_v' in list(self.model_ginitime.columns):
                    data_df=data_df.rename({'auc_v':"vaild AUC",
                                            'count_v':"vaild样本数",
                                            'bad_rate_v':"vaild坏账率"},axis=1)
                    col = col + ["vaild AUC", "vaild样本数", "vaild坏账率"]
                if 'count_oot' in list(self.model_ginitime.columns):
                    data_df=data_df.rename({
                                            'auc_oot':"OOT AUC",
                                            'count_oot':"OOT样本数",
                                            'bad_rate_oot':"OOT坏账率"},axis=1)
                    col = col + ["OOT AUC", "OOT样本数", "OOT坏账率"]
                if 'count_reject' in list(self.model_ginitime.columns):
                    data_df=data_df.rename({
                                            'auc_reject':"Reject AUC",
                                            'count_reject':"Reject样本数",
                                            'bad_rate_reject':"Reject坏账率"},axis=1)
                    col = col + ["Reject AUC", "Reject样本数", "Reject坏账率"]
                rows = dataframe_to_rows(data_df[col], index=False)
                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        workSheet3.cell(row=r_idx + 3, column=c_idx, value=value)
                        if r_idx == 1:
                            boder_set(workSheet3, start_col=c_idx, end_col=c_idx, start_row=r_idx + 3,
                                      end_row=r_idx + 3, tile=True)

                # 模型表现稳定性
                workSheet3.cell(1, len(col)+4).value = '模型表现稳定性'
                boder_set(workSheet3, start_col=len(col)+4, end_col=len(col)+4, start_row=1, end_row=1, tile=1)
                imgdata = io.BytesIO()
                self.gg_time.savefig(imgdata, format='png')
                img = openpyxl.drawing.image.Image(imgdata)
                img.anchor = '%s3'%(col_num_alp(len(col)+4))
                workSheet3.add_image(img)
        #-------------------------------------------------------------------------------------------
            # 模型表现稳定性
            workSheet3.cell(len(self.model_ginitime)+6, 1).value = '分数分布稳定性'
            boder_set(workSheet3, start_col=1, end_col=1, start_row=len(self.model_ginitime)+6, end_row=len(self.model_ginitime)+6, tile=1)
            imgdata = io.BytesIO()
            self.gg_dis.savefig(imgdata, format='png')
            img = openpyxl.drawing.image.Image(imgdata)
            img.anchor = 'A%d' % (len(self.model_ginitime)+8)
            workSheet3.add_image(img)
        workSheet4 = wb.create_sheet('模型变量')
        row_num=0
        var_list = list(set(self.scorecarddf['variable_name']))
        try:
            var_list.remove('const')
        except:
            pass
        if self.predict_vaild_data.empty==False:
            flag_v=True
        else:
            flag_v=False
        for var_label in var_list:
            if var_label != 'const':
                if list(self.scorecarddf[self.scorecarddf['variable_name'] == var_label]['var_type'])[0] == 'add':
                    report = self.scorecarddf[self.scorecarddf['variable_name'] == var_label][
                        ['f_group', 'label', 'f_Bad_rate', 'f_N_obs']]
                    report['iv'] = None
                    report['miss_rate'] = 0
                else:
                    report = self.f_group_report[self.f_group_report['variable_name'] == var_label]
                tree_rep = report[['f_group', 'label', 'iv', 'f_Bad_rate', 'f_N_obs', 'miss_rate']].rename({
                                        'f_group': "组数",
                                        'label': "标识",
                                        'iv': "信息熵",
                                        'f_Bad_rate': "组坏账率",
                                        'f_N_obs': "组坏账数",
                                        'miss_rate': "缺失率"}, axis=1)
                workSheet4.cell(row_num+1, 1).value = var_label
                boder_set(workSheet4, start_col=1, end_col=1, start_row=row_num+1,
                          end_row=row_num+1, tile=1)
                workSheet4.cell(row_num+3, 1).value = '变量分组数据'
                boder_set(workSheet4, start_col=1, end_col=1, start_row=row_num+3,
                          end_row=row_num+3, tile=1)
                rows = dataframe_to_rows(tree_rep, index=False)
                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        workSheet4.cell(row=r_idx + row_num+4, column=c_idx , value=value)
                        if r_idx == 1:
                            boder_set(workSheet4, start_col=c_idx ,
                                      end_col=c_idx , start_row= r_idx + row_num+4,
                                      end_row=r_idx + row_num+4, tile=True)
                        else:
                            boder_set(workSheet4, start_col=c_idx ,
                                      end_col=c_idx , start_row= r_idx + row_num+4,
                                      end_row=r_idx + row_num+4)
                if list(self.scorecarddf[self.scorecarddf['variable_name'] == var_label]['var_type'])[0] == 'add':
                    if self.model_var_type == 'WOE':
                        model_ginitime = self.calculateginitime(var_label, self.predict_train_data, self.train_time_id,self.train_target,flag_v,self.predict_vaild_data,
                                                                self.predict_oot_data,self.oot_target,self.oot_time_id,self.predict_reject_data,self.reject_time_id,self.reject_target)
                    else:
                        model_ginitime = self.calculateginitime('f_group_%s' % var_label, self.predict_train_data, self.train_time_id,self.train_target,flag_v,self.predict_vaild_data,
                                                                self.predict_oot_data,self.oot_target,self.oot_time_id,self.predict_reject_data,self.reject_time_id,self.reject_target)
                else:
                    model_ginitime = self.calculateginitime('woe_%s' % var_label, self.predict_train_data,
                                                            self.train_time_id, self.train_target, flag_v,
                                                            self.predict_vaild_data,
                                                            self.predict_oot_data, self.oot_target,
                                                            self.oot_time_id, self.predict_reject_data,
                                                            self.reject_time_id, self.reject_target)
                if model_ginitime.empty==False:
                    col = ["时间ID"]
                    data_df = model_ginitime
                    data_df = data_df.rename({'timeid': "时间ID"}, axis=1)
                    if 'count' in list(self.model_ginitime.columns):
                        data_df = data_df.rename({
                            'auc': "Trian AUC",
                            'count': "Trian样本数",
                            'bad_rate': "Trian坏账率"}, axis=1)
                        col = col + ["Trian AUC", "Trian样本数", "Trian坏账率"]
                    if 'count_v' in list(self.model_ginitime.columns):
                        data_df = data_df.rename({'auc_v': "vaild AUC",
                                                  'count_v': "vaild样本数",
                                                  'bad_rate_v': "vaild坏账率"}, axis=1)
                        col = col + ["vaild AUC", "vaild样本数", "vaild坏账率"]
                    if 'count_oot' in list(self.model_ginitime.columns):
                        data_df = data_df.rename({
                            'auc_oot': "OOT AUC",
                            'count_oot': "OOT样本数",
                            'bad_rate_oot': "OOT坏账率"}, axis=1)
                        col = col + ["OOT AUC", "OOT样本数", "OOT坏账率"]
                    if 'count_reject' in list(self.model_ginitime.columns):
                        data_df = data_df.rename({
                            'auc_reject': "Reject AUC",
                            'count_reject': "Reject样本数",
                            'bad_rate_reject': "Reject坏账率"}, axis=1)
                        col = col + ["Reject AUC", "Reject样本数", "Reject坏账率"]
                    workSheet4.cell(row_num + 3, len(tree_rep.columns)+3).value = '变量表现稳定性'
                    boder_set(workSheet4, start_col=len(tree_rep.columns)+3, end_col=len(tree_rep.columns)+3, start_row=row_num + 3,
                              end_row=row_num + 3, tile=1)
                    rows = dataframe_to_rows(data_df[col], index=False)
                    for r_idx, row in enumerate(rows, 1):
                        for c_idx, value in enumerate(row, 1):
                            workSheet4.cell(row=row_num+r_idx + 4, column=len(tree_rep.columns)+2+c_idx, value=value)
                            if r_idx == 1:
                                boder_set(workSheet4, start_col=len(tree_rep.columns)+2+c_idx, end_col=len(tree_rep.columns)+2+c_idx, start_row=r_idx + 4+row_num,
                                          end_row=r_idx + 4+row_num, tile=True)
                            else:
                                boder_set(workSheet4, start_col=len(tree_rep.columns)+2+c_idx, end_col=len(tree_rep.columns)+2+c_idx, start_row=r_idx + 4+row_num,
                                          end_row=r_idx + 4+row_num)
                ccc = self.groupplot_datainit(modify_var=var_label)
                re_fig = self.plot_group(bad_pct=ccc[0], total_pct=ccc[1], final=ccc[2],directtion='H')
                # 模型表现稳定性
                workSheet4.cell(row_num + 3, len(tree_rep.columns)+len(model_ginitime.columns)+5).value = '变量表现稳定性'
                boder_set(workSheet4, start_col=len(tree_rep.columns)+len(model_ginitime.columns)+5, end_col=len(tree_rep.columns)+len(model_ginitime.columns)+5,start_row=row_num + 3, end_row=row_num + 3, tile=1)
                imgdata = io.BytesIO()
                re_fig.savefig(imgdata, format='png')
                img = openpyxl.drawing.image.Image(imgdata)
                img.anchor = '%s%d' % (col_num_alp(len(tree_rep.columns)+len(model_ginitime.columns)+5),row_num + 5)
                workSheet4.add_image(img)
                row_num=row_num+max(30,len(tree_rep),len(model_ginitime))
        if self.lasso_df.empty==False:
            workSheet5 = wb.create_sheet('LASSO')
            workSheet5.cell(1, 1).value = 'Lasso数据集'
            boder_set(workSheet5, start_col=1, end_col=1,start_row=1,end_row=1, tile=1)
            tree_lasso=self.step_show_df.copy()
            col=["变量名","步数","训练集LLR","训练集AIC","训练集BIC"]
            tree_lasso=tree_lasso.rename({'variable_name':"变量名",
                                            'step':"步数",
                                            'llr':"训练集LLR",
                                            'aic':"训练集AIC",
                                            'bic':"训练集BIC"},axis=1)
            if ('aic_v' in tree_lasso.columns) and ('bic_v' in tree_lasso.columns):
                tree_lasso = tree_lasso.rename({'llr_v':"验证集LLR",
                                                'aic_v':"验证集AIC",
                                                'bic_v':"验证集BIC"}, axis=1)
                col=col+["验证集LLR","验证集AIC","验证集BIC"]
            rows = dataframe_to_rows(tree_lasso[col], index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    workSheet5.cell(row= r_idx + 2, column= c_idx, value=value)
                    if r_idx == 1:
                        boder_set(workSheet4, start_col=c_idx,
                                  end_col=c_idx, start_row=r_idx + 2,
                                  end_row=r_idx + 2, tile=True)
                    else:
                        boder_set(workSheet4, start_col= c_idx,
                                  end_col=c_idx, start_row=r_idx + 2,
                                  end_row=r_idx + 2)

            workSheet5.cell( 3,len(col) + 6).value = 'Lasso图表'
            boder_set(workSheet5, start_col=len(col) + 6, end_col=len(col) + 6, start_row=3,
                      end_row=3, tile=1)
            imgdata = io.BytesIO()
            self.lasso_fig.savefig(imgdata, format='png')
            img = openpyxl.drawing.image.Image(imgdata)
            img.anchor = '%s%d' %(col_num_alp(len(col)+6),3)
            workSheet5.add_image(img)
        workSheet6 = wb.create_sheet('相关矩阵')
        workSheet6.cell(1, 1).value = '相关矩阵'
        rows = dataframe_to_rows(self.corr_data, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                workSheet6.cell(row=r_idx + 2, column=c_idx+2, value=value)
                if r_idx == 1:
                    boder_set(workSheet6, start_col=c_idx+2,
                              end_col=c_idx+2, start_row=r_idx + 2,
                              end_row=r_idx + 2, tile=True)
                else:
                    boder_set(workSheet6, start_col=c_idx+2,
                              end_col=c_idx+2, start_row=r_idx + 2,
                              end_row=r_idx + 2)
            if r_idx!=1:
                workSheet6.cell(row=r_idx + 2, column=2, value=list(self.corr_data.columns)[r_idx-2])
                boder_set(workSheet6, start_col=2,
                          end_col= 2, start_row=r_idx + 2,
                          end_row=r_idx + 2, tile=True)
        workSheet6.conditional_formatting.add(
                                        'C4:%s%d'%(col_num_alp(len(self.corr_data.columns)+2),3+len(self.corr_data)),
                                        ColorScaleRule(start_type='min', start_value=-1, start_color='FFF9CC',
                                                       mid_type='num', mid_value=0, mid_color='FFFFFF',
                                                          end_type='max', end_value=1, end_color='DB3232') )
        wb.save(self.project_path + '/' + '%s_model_report.xlsx' % self.node_name)